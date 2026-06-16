# =============================================================================
# BioBlio — QGIS
# Auteur : généré via assistant IA
# Usage  : coller dans la console Python de QGIS (Menu Extensions → Console Python)
#          ou lancer via : exec(open(r"chemin\vers\ce\fichier.py").read())
# =============================================================================

import requests
import processing
from qgis.PyQt.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QComboBox, QPushButton, QCheckBox, QScrollArea,
    QWidget, QMessageBox, QProgressBar, QFrame, QApplication,
    QCompleter  
)
from qgis.PyQt.QtCore import Qt, QStringListModel, QDate, QDateTime
from qgis.PyQt.QtGui import QFont, QColor, QPalette, QPixmap
from qgis.core import (
    QgsVectorLayer, QgsProject,
    QgsLayerTreeGroup, QgsLayerTreeLayer,
    QgsLinePatternFillSymbolLayer,
    QgsFillSymbol, QgsLineSymbol, QgsSingleSymbolRenderer,
    QgsSimpleFillSymbolLayer, QgsSimpleLineSymbolLayer,
    QgsUnitTypes
)
from qgis.utils import iface

WFS_GPF  = "https://data.geopf.fr/wfs/ows"
WFS_SAND = "https://services.sandre.eaufrance.fr/geo/sandre"
WFS_RPDZH = "http://wms.reseau-zones-humides.org/cgi-bin/wmsfma"
WFS_GEOPORTAIL = "https://wxs.ign.fr/geoportail/wfs"
GBIF_API = "https://api.gbif.org/v1/occurrence/search"
MM = QgsUnitTypes.RenderMillimeters
API_CARTO_CADASTRE  = "https://apicarto.ign.fr/api/cadastre/parcelle"
API_CARTO_BATIMENT  = "https://apicarto.ign.fr/api/cadastre/batiment"

# Groupes taxonomiques GBIF (taxonKey)
GBIF_GROUPES = [
    ("Tous groupes",    None),
    ("🦋 Insectes",     216),
    ("🌿 Plantes",      6),
    ("🐦 Oiseaux",      212),
    ("🦇 Chiroptères",  734),
    ("🐾 Mammifères",   359),
    ("🐸 Amphibiens",   131),
    ("🦎 Reptiles",     358),
    ("🐟 Poissons",     204),
    ("🦀 Crustacés",    229),
    ("🐌 Mollusques",   52),
    ("🍄 Champignons",  5),
    ("🌾 Bryophytes",   35),
    ("🦟 Arachnides",   367),
]

# (Catégorie, Nom affiché, typename WFS, serveur, fill_rgba, stroke_hex)
CATALOGUE = [
    ("🛡 Zonages réglementaires", "🔴 ZNIEFF Type 1",
     "patrinat_znieff1:znieff1",        WFS_GPF,  "255,130,0,100",  "#CC5000"),
    ("🛡 Zonages réglementaires", "🟠 ZNIEFF Type 2",
     "patrinat_znieff2:znieff2",        WFS_GPF,  "240,200,60,90",  "#9A7A00"),
    ("🛡 Zonages réglementaires", "🟢 Natura 2000 — ZSC (Habitats)",
     "patrinat_sic:sic",                WFS_GPF,  "34,139,34,100",  "#006400"),
    ("🛡 Zonages réglementaires", "🔵 Natura 2000 — ZPS (Oiseaux)",
     "patrinat_zps:zps",                WFS_GPF,  "30,100,210,110", "#002299"),
    ("🛡 Zonages réglementaires", "🟣 Réserves Naturelles Nationales",
     "patrinat_rnn:rnn",                WFS_GPF,  "180,0,200,90",   "#800090"),
    ("🛡 Zonages réglementaires", "🪻 Réserves Naturelles Régionales",
     "patrinat_rnr:rnr",                WFS_GPF,  "200,50,200,90",  "#900070"),
    ("🛡 Zonages réglementaires", "🌲 Parcs Naturels Régionaux",
     "patrinat_pnr:pnr",                WFS_GPF,  "100,180,100,80", "#2A7A2A"),
    ("🛡 Zonages réglementaires", "🌿 Arrêtés Protection Biotope",
     "patrinat_apb:apb",                WFS_GPF,  "255,200,50,90",  "#AA7000"),
    ("🛡 Zonages réglementaires", "🍃 Conservatoires Espaces Naturels",
     "patrinat_cen:cen",                WFS_GPF,  "50,200,150,90",  "#007755"),
    ("💧 Données eau", "🔷 Cours d'eau L214-17 — Liste 1",
     "sa:SegClassContinuiteEcoListe1_FXX", WFS_SAND, None,          "#08519C"),
    ("💧 Données eau", "🔹 Cours d'eau L214-17 — Liste 2",
     "sa:SegClassContinuiteEcoListe2_FXX", WFS_SAND, None,          "#3182BD"),
    ("💧 Données eau", "🌊 SAGE",
     "sa:Sage_FXX",                     WFS_SAND, "100,180,230,55", "#0A4FA8"),
    ("💧 Zones humides", "💧 Zones humides d'importance majeure — RPDZH",
     "ZHIM",                            WFS_RPDZH, "50,160,220,80", "#006699"),

    ("💧 Zones humides", "🦆 Sites Ramsar — périmètres",
     "patrinat_ramsar:pnm",             WFS_GPF, "30,120,200,70", "#004C99"),
    ("📐 Limites administratives", "🏘 Communes",
     "LIMITES_ADMINISTRATIVES_EXPRESS.LATEST:commune",     WFS_GPF, "255,255,255,0", "#888888"),
    ("📐 Limites administratives", "🗺 Départements",
     "LIMITES_ADMINISTRATIVES_EXPRESS.LATEST:departement", WFS_GPF, "255,255,255,0", "#444444"),
    ("📐 Limites administratives", "📍 Régions",
     "LIMITES_ADMINISTRATIVES_EXPRESS.LATEST:region",      WFS_GPF, "255,255,255,0", "#222222"),

    ("🏙 Urbanisme (GPU)", "🟦 Zones PLU / Cartes communales",
     "wfs_du:zone_urba|COMMUNE_ONLY", WFS_GPF, "30,144,255,50", "#0050AA"),
    ("🏙 Urbanisme (GPU)", "🌲 Espaces Boisés Classés (EBC)",
     "wfs_du:prescription_surf|typepsc=01|COMMUNE_ONLY",
                                   WFS_GPF, "34,139,34,70",   "#005000"),
    ("🏙 Urbanisme (GPU)", "🌿 Éléments paysagers à préserver",
     "wfs_du:prescription_surf|typepsc=07|stypepsc=00,02,04|COMMUNE_ONLY",
                                   WFS_GPF, "144,238,144,70", "#2E8B00"),
    ("🏙 Urbanisme (GPU)", "🌳 Arbres / éléments naturels ponctuels à préserver",
     "wfs_du:prescription_pct|typepsc=07|stypepsc=04,05|COMMUNE_ONLY",
                                   WFS_GPF, None, "#1B7F3A"),

    ("🏛 Patrimoine culturel (SUP)", "🏛 Monuments historiques — AC1",
     "wfs_sup:assiette_sup_s|suptype=ac1",
                                   WFS_GPF, "255,215,0,45", "#B8860B"),
    ("🏛 Patrimoine culturel (SUP)", "🏞 Sites classés / inscrits — AC2",
     "wfs_sup:assiette_sup_s|suptype=ac2",
                                   WFS_GPF, "180,80,200,45", "#7A3E9D"),
    ("🏛 Patrimoine culturel (SUP)", "🏘 Sites patrimoniaux remarquables — AC4",
     "wfs_sup:assiette_sup_s|suptype=ac4",
                                   WFS_GPF, "200,80,80,45", "#8B0000"),

    ("🧾 Cadastre", "🧾 Parcelles cadastrales",
     "CADASTRALPARCELS.PARCELLAIRE_EXPRESS:parcelle",
                                   WFS_GPF, "255,255,255,0", "#CC00CC"),

    ("🧾 Cadastre", "🏠 Bâtiments cadastraux",
     "CADASTRALPARCELS.PARCELLAIRE_EXPRESS:batiment",
                                   WFS_GPF, "160,160,160,90", "#555555"),
]

def apply_style(lyr, fill_rgba, stroke_hex, width=0.4):
    """Symbologie identique au projet existant selon le nom de la couche."""
    nom = lyr.name()
    MM  = QgsUnitTypes.RenderMillimeters

    from qgis.core import (QgsLinePatternFillSymbolLayer,
                            QgsSimpleFillSymbolLayer, QgsSimpleLineSymbolLayer)
    from qgis.PyQt.QtCore import Qt

    # ── ZNIEFF Type 1 : contour rouge foncé 0.55mm, fond transparent
    if "znieff" in nom.lower() and ("type 1" in nom.lower() or "znieff1" in nom.lower()):
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)
        fl = QgsSimpleFillSymbolLayer()
        fl.setBrushStyle(Qt.NoBrush)
        fl.setStrokeColor(QColor("#AA1100"))
        fl.setStrokeWidth(0.55); fl.setStrokeWidthUnit(MM)
        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    # ── ZNIEFF Type 2 : hachures croisées orange 45°+135°, contour 0.40mm
    elif "znieff" in nom.lower() and ("type 2" in nom.lower() or "znieff2" in nom.lower()):
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)
        for angle in [45.0, 135.0]:
            h = QgsLinePatternFillSymbolLayer()
            h.setColor(QColor("#CC6600"))
            h.setLineWidth(0.20); h.setLineWidthUnit(MM)
            h.setLineAngle(angle)
            h.setDistance(2.5); h.setDistanceUnit(MM)
            sym.appendSymbolLayer(h)
        fl = QgsSimpleFillSymbolLayer()
        fl.setBrushStyle(Qt.NoBrush)
        fl.setStrokeColor(QColor("#884400"))
        fl.setStrokeWidth(0.40); fl.setStrokeWidthUnit(MM)
        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    # ── ZSC : fond vert semi-transparent, contour vert foncé 0.35mm
    elif "zsc" in nom.lower() or "sic" in nom.lower() or "habitats" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)
        fl = QgsSimpleFillSymbolLayer()
        fl.setFillColor(QColor(34, 139, 34, 70))
        fl.setStrokeColor(QColor("#005000"))
        fl.setStrokeWidth(0.35); fl.setStrokeWidthUnit(MM)
        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    # ── ZPS : hachures bleues 45°, contour bleu marine 0.50mm
    elif "zps" in nom.lower() or "oiseaux" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)
        h = QgsLinePatternFillSymbolLayer()
        h.setColor(QColor("#1050CC"))
        h.setLineWidth(0.20); h.setLineWidthUnit(MM)
        h.setLineAngle(45.0)
        h.setDistance(2.0); h.setDistanceUnit(MM)
        sym.appendSymbolLayer(h)
        fl = QgsSimpleFillSymbolLayer()
        fl.setBrushStyle(Qt.NoBrush)
        fl.setStrokeColor(QColor("#002299"))
        fl.setStrokeWidth(0.50); fl.setStrokeWidthUnit(MM)
        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    # ── SAGE : hachures horizontales bleues 0°, contour bleu marine 0.50mm
    elif "sage" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)
        h = QgsLinePatternFillSymbolLayer()
        h.setColor(QColor("#2171B5"))
        h.setLineWidth(0.20); h.setLineWidthUnit(MM)
        h.setLineAngle(0.0)
        h.setDistance(2.5); h.setDistanceUnit(MM)
        sym.appendSymbolLayer(h)
        fl = QgsSimpleFillSymbolLayer()
        fl.setBrushStyle(Qt.NoBrush)
        fl.setStrokeColor(QColor("#08306B"))
        fl.setStrokeWidth(0.50); fl.setStrokeWidthUnit(MM)
        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    # ── L214 Liste 1 : ligne bleue marine pleine 0.60mm
    elif "liste1" in nom.lower() or "liste 1" in nom.lower():
        sym = QgsLineSymbol()
        sym.deleteSymbolLayer(0)
        sl = QgsSimpleLineSymbolLayer()
        sl.setColor(QColor("#08519C"))
        sl.setWidth(0.60); sl.setWidthUnit(MM)
        sl.setPenStyle(Qt.SolidLine)
        sym.appendSymbolLayer(sl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    # ── L214 Liste 2 : ligne bleue ciel tirets 0.45mm
    elif "liste2" in nom.lower() or "liste 2" in nom.lower():
        sym = QgsLineSymbol()
        sym.deleteSymbolLayer(0)
        sl = QgsSimpleLineSymbolLayer()
        sl.setColor(QColor("#3182BD"))
        sl.setWidth(0.45); sl.setWidthUnit(MM)
        sl.setPenStyle(Qt.DashLine)
        sym.appendSymbolLayer(sl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    # ── Zones humides RPDZH : bleu humide semi-transparent
    elif "zones humides" in nom.lower() or "rpdzh" in nom.lower() or "zhim" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)

        fl = QgsSimpleFillSymbolLayer()
        fl.setFillColor(QColor(50, 160, 220, 80))
        fl.setStrokeColor(QColor("#006699"))
        fl.setStrokeWidth(0.35)
        fl.setStrokeWidthUnit(MM)

        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

        # ── Sites Ramsar : bleu soutenu semi-transparent
    elif "ramsar" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)

        fl = QgsSimpleFillSymbolLayer()
        fl.setFillColor(QColor(30, 120, 200, 70))
        fl.setStrokeColor(QColor("#004C99"))
        fl.setStrokeWidth(0.45)
        fl.setStrokeWidthUnit(MM)

        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    elif "parcelles cadastrales" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)

        fl = QgsSimpleFillSymbolLayer()
        fl.setBrushStyle(Qt.NoBrush)
        fl.setStrokeColor(QColor("#252425D5"))
        fl.setStrokeWidth(0.18)
        fl.setStrokeWidthUnit(MM)

        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    elif "bâtiments cadastraux" in nom.lower() or "batiments cadastraux" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)

        fl = QgsSimpleFillSymbolLayer()
        fl.setFillColor(QColor(160, 160, 160, 90))
        fl.setStrokeColor(QColor("#555555"))
        fl.setStrokeWidth(0.15)
        fl.setStrokeWidthUnit(MM)

        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    elif "monuments historiques" in nom.lower() or "ac1" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)

        h = QgsLinePatternFillSymbolLayer()
        h.setColor(QColor("#B8860B"))
        h.setLineWidth(0.18)
        h.setLineWidthUnit(MM)
        h.setLineAngle(45.0)
        h.setDistance(2.2)
        h.setDistanceUnit(MM)
        sym.appendSymbolLayer(h)

        fl = QgsSimpleFillSymbolLayer()
        fl.setFillColor(QColor(255, 215, 0, 35))
        fl.setStrokeColor(QColor("#8B6508"))
        fl.setStrokeWidth(0.45)
        fl.setStrokeWidthUnit(MM)
        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    elif "sites classés" in nom.lower() or "sites classes" in nom.lower() or "ac2" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)

        fl = QgsSimpleFillSymbolLayer()
        fl.setFillColor(QColor(180, 80, 200, 45))
        fl.setStrokeColor(QColor("#7A3E9D"))
        fl.setStrokeWidth(0.45)
        fl.setStrokeWidthUnit(MM)
        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    elif "sites patrimoniaux remarquables" in nom.lower() or "ac4" in nom.lower():
        sym = QgsFillSymbol()
        sym.deleteSymbolLayer(0)

        h = QgsLinePatternFillSymbolLayer()
        h.setColor(QColor("#8B0000"))
        h.setLineWidth(0.16)
        h.setLineWidthUnit(MM)
        h.setLineAngle(135.0)
        h.setDistance(2.4)
        h.setDistanceUnit(MM)
        sym.appendSymbolLayer(h)

        fl = QgsSimpleFillSymbolLayer()
        fl.setFillColor(QColor(200, 80, 80, 35))
        fl.setStrokeColor(QColor("#8B0000"))
        fl.setStrokeWidth(0.45)
        fl.setStrokeWidthUnit(MM)
        sym.appendSymbolLayer(fl)
        lyr.setRenderer(QgsSingleSymbolRenderer(sym))

    # ── Autres (RNN, RNR, PNR, APB, CEN, communes, depts, régions)
    # → utilise la symbologie générique du catalogue (fill_rgba + stroke_hex)
    else:
        if lyr.geometryType() == 2:
            sym = QgsFillSymbol()
            sym.deleteSymbolLayer(0)
            fl = QgsSimpleFillSymbolLayer()
            if fill_rgba:
                r, g, b, a = [int(x) for x in fill_rgba.split(',')]
                fl.setFillColor(QColor(r, g, b, a))
            else:
                fl.setBrushStyle(Qt.NoBrush)
            fl.setStrokeColor(QColor(stroke_hex))
            fl.setStrokeWidth(width); fl.setStrokeWidthUnit(MM)
            sym.appendSymbolLayer(fl)
            lyr.setRenderer(QgsSingleSymbolRenderer(sym))
        elif lyr.geometryType() == 1:
            sym = QgsLineSymbol()
            sym.deleteSymbolLayer(0)
            sl = QgsSimpleLineSymbolLayer()
            sl.setColor(QColor(stroke_hex))
            sl.setWidth(0.5); sl.setWidthUnit(MM)
            sym.appendSymbolLayer(sl)
            lyr.setRenderer(QgsSingleSymbolRenderer(sym))
        elif lyr.geometryType() == 0:
            from qgis.core import QgsMarkerSymbol, QgsSimpleMarkerSymbolLayer

            sym = QgsMarkerSymbol()
            sym.deleteSymbolLayer(0)

            mk = QgsSimpleMarkerSymbolLayer()
            mk.setShape(QgsSimpleMarkerSymbolLayer.Shape.Circle)
            mk.setSize(2.2)
            mk.setSizeUnit(MM)
            mk.setColor(QColor("#2ECC71"))
            mk.setStrokeColor(QColor(stroke_hex))
            mk.setStrokeWidth(0.25)
            mk.setStrokeWidthUnit(MM)

            sym.appendSymbolLayer(mk)
            sym.setOpacity(0.85)
            lyr.setRenderer(QgsSingleSymbolRenderer(sym))

def build_wfs_uri(serveur, typename_raw, bbox_str=None):
    """Construit l'URI WFS et retourne (uri, filtre_expression, is_gpu).
    Format : 'typename' ou 'typename|champ=val|GPU'
    is_gpu=True signale que la couche GPU nécessite un filtrage par code INSEE
    """
    parts = typename_raw.split("|")
    typename = parts[0]
    is_gpu = "GPU" in parts or "COMMUNE_ONLY" in parts

    filtres = []
    filtres_ogc = []
    for part in parts[1:]:
        # Ignorer tous les marqueurs sans '=' (GPU, COMMUNE_ONLY, etc.)
        if '=' not in part:
            continue
        champ, vals = part.split("=", 1)
        val_list = vals.split(",")
        if len(val_list) == 1:
            filtres.append(f'"{champ}" = \'{val_list[0]}\'')
            filtres_ogc.append(
                f"<PropertyIsEqualTo>"
                f"<PropertyName>{champ}</PropertyName>"
                f"<Literal>{val_list[0]}</Literal>"
                f"</PropertyIsEqualTo>"
            )
        else:
            in_vals = ", ".join(f"'{v}'" for v in val_list)
            filtres.append(f'"{champ}" IN ({in_vals})')
            conditions = "".join(
                f"<PropertyIsEqualTo>"
                f"<PropertyName>{champ}</PropertyName>"
                f"<Literal>{v}</Literal>"
                f"</PropertyIsEqualTo>"
                for v in val_list
            )
            filtres_ogc.append(f"<Or>{conditions}</Or>")

    filtre_expr = " AND ".join(filtres)

    # Les WFS SUP du GPU refusent BBOX + FILTER séparés.
    # On pousse donc le filtre attributaire et spatial dans un seul filtre OGC.
    if typename.startswith("wfs_sup:") and filtres_ogc:
        conditions = list(filtres_ogc)
        if bbox_str:
            xmin, ymin, xmax, ymax = bbox_str.split(",")[:4]
            conditions.append(
                f"<BBOX>"
                f"<PropertyName>the_geom</PropertyName>"
                f"<gml:Envelope xmlns:gml=\"http://www.opengis.net/gml\" "
                f"srsName=\"EPSG:4326\">"
                f"<gml:lowerCorner>{xmin} {ymin}</gml:lowerCorner>"
                f"<gml:upperCorner>{xmax} {ymax}</gml:upperCorner>"
                f"</gml:Envelope>"
                f"</BBOX>"
            )

        if len(conditions) == 1:
            filtre_ogc = f"<Filter>{conditions[0]}</Filter>"
        else:
            filtre_ogc = f"<Filter><And>{''.join(conditions)}</And></Filter>"

        uri = (f"url='{serveur}' typename='{typename}' version='1.1.0' "
               f"srsname='EPSG:4326' filter='{filtre_ogc}'")
        return uri, "", is_gpu

    if bbox_str:
        uri = (f"url='{serveur}' typename='{typename}' version='auto' "
               f"srsname='EPSG:4326' restrictToRequestBBOX='1' "
               f"bbox='{bbox_str},EPSG:4326'")
    else:
        uri = (f"url='{serveur}' typename='{typename}' version='auto' "
               f"srsname='EPSG:4326' restrictToRequestBBOX='0'")

    return uri, filtre_expr, is_gpu

def layer_has_features(lyr):
    """Teste rapidement si une couche contient au moins une entité."""
    if not lyr or not lyr.isValid():
        return False
    try:
        return next(lyr.getFeatures(), None) is not None
    except Exception:
        return False

def is_sup_layer(typename_raw):
    """Repère les couches de servitudes GPU, dont les géométries sont parfois fragiles."""
    return typename_raw.split("|")[0].startswith("wfs_sup:")

def materialize_layer(lyr, nom=None):
    """Copie une couche distante en mémoire pour éviter les relances WFS pendant les traitements."""
    if not lyr or not lyr.isValid():
        return lyr

    try:
        from qgis.core import QgsWkbTypes

        geom_name = QgsWkbTypes.displayString(lyr.wkbType())
        if not geom_name or geom_name == "Unknown":
            if lyr.geometryType() == 0:
                geom_name = "Point"
            elif lyr.geometryType() == 1:
                geom_name = "LineString"
            else:
                geom_name = "Polygon"

        crs_id = lyr.crs().authid() or "EPSG:4326"
        lyr_mem = QgsVectorLayer(f"{geom_name}?crs={crs_id}", nom or lyr.name(), "memory")
        pr_mem = lyr_mem.dataProvider()
        pr_mem.addAttributes(lyr.fields().toList())
        lyr_mem.updateFields()
        pr_mem.addFeatures(list(lyr.getFeatures()))
        try:
            pr_mem.createSpatialIndex()
        except Exception:
            pass
        lyr_mem.updateExtents()
        return lyr_mem
    except Exception:
        return lyr

def exclude_from_excel(nom):
    """Indique si une couche doit être exclue des exports Excel."""
    nom_lower = nom.lower()
    return (
        "parcelles cadastrales" in nom_lower
        or "bâtiments cadastraux" in nom_lower
        or "batiments cadastraux" in nom_lower
    )

def load_cadastre_api(endpoint, lyr_territoire, nom, lbl_status, app, code_insee=None):
    """Charge des entités cadastrales (parcelles ou bâtiments) via API Carto.
    PAGE_SIZE = 5000 → la plupart des communes tiennent en 1 seule requête.
    Les pages supplémentaires sont récupérées en parallèle (MAX_WORKERS threads).
    """
    import json
    import os
    import tempfile
    import requests
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from qgis.core import QgsGeometry, QgsVectorLayer

    PAGE_SIZE   = 5000   # API Carto supporte jusqu'à 5000 entités par page
    MAX_WORKERS = 8

    def build_geom_params():
        if not lyr_territoire:
            return None
        geoms = [f.geometry() for f in lyr_territoire.getFeatures() if f.hasGeometry()]
        if not geoms:
            return None
        geom = QgsGeometry.unaryUnion(geoms)
        geom = geom.simplify(0.00002)
        geom_json = json.loads(geom.asJson(6))
        return {"geom": json.dumps(geom_json, separators=(",", ":"))}

    def fetch_page(base_params, start):
        params = dict(base_params)
        params["_limit"] = PAGE_SIZE
        if start > 0:
            params["_start"] = start
        try:
            r = requests.get(endpoint, params=params, timeout=90)
            if r.status_code == 200:
                return start, r.json().get("features", [])
            return start, None
        except Exception:
            return start, []

    # Priorité : code INSEE (commune) puis géométrie (fallback)
    param_sets = []
    if code_insee:
        param_sets.append((f"commune {code_insee}", {"code_insee": code_insee}))
    geom_params = build_geom_params()
    if geom_params:
        param_sets.append(("zone", geom_params))

    if not param_sets:
        return None

    all_features = []

    for mode_label, base_params in param_sets:
        # ── Page 0 : vérifie que des données existent ─────────────────────
        lbl_status.setText(f"Cadastre {mode_label} — téléchargement...")
        app.processEvents()

        _, first_page = fetch_page(base_params, 0)
        if not first_page:
            continue  # pas de données ou erreur HTTP → fallback suivant

        all_features = list(first_page)

        if len(first_page) < PAGE_SIZE:
            break  # tout récupéré en une seule requête (cas habituel)

        # ── Pages suivantes : toutes en parallèle ─────────────────────────
        offset = PAGE_SIZE
        while True:
            starts = [offset + i * PAGE_SIZE for i in range(MAX_WORKERS)]
            lbl_status.setText(
                f"Cadastre {mode_label} — {len(all_features)} entités, "
                f"suite en cours...")
            app.processEvents()

            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
                futures = {ex.submit(fetch_page, base_params, s): s for s in starts}
                results = {}
                for fut in as_completed(futures):
                    s, feats = fut.result()
                    results[s] = feats if feats is not None else []

            done = False
            for s in sorted(starts):
                feats = results.get(s, [])
                all_features.extend(feats)
                if len(feats) < PAGE_SIZE:
                    done = True
                    break
            if done:
                break
            offset += MAX_WORKERS * PAGE_SIZE

        break  # succès

    if not all_features:
        return None

    fc = {"type": "FeatureCollection", "features": all_features}
    fd, path = tempfile.mkstemp(prefix="cadastre_", suffix=".geojson")
    os.close(fd)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(fc, f)

    lyr = QgsVectorLayer(path, nom, "ogr")
    return lyr if lyr.isValid() else None


def load_cadastre_parcelles_api(lyr_territoire, nom, lbl_status, app, code_insee=None):
    """Alias conservé pour compatibilité — délègue à load_cadastre_api."""
    return load_cadastre_api(
        API_CARTO_CADASTRE, lyr_territoire, nom, lbl_status, app,
        code_insee=code_insee
    )

def load_gpu_layer(typename_raw, filtre_expr, lyr_territoire, nom, lbl_status, app):
    """Charge une couche GPU via filtre OGC XML par partition (code INSEE commune)."""
    from qgis.core import QgsFeatureRequest, QgsExpression

    WFS_GPF = "https://data.geopf.fr/wfs/ows"
    parts    = typename_raw.split("|")
    typename = parts[0]

    # Étape 1 — codes INSEE des communes du territoire
    lbl_status.setText(f"GPU — communes pour {nom}...")
    app.processEvents()

    ext = lyr_territoire.extent()
    mx  = ext.width()  * 0.02
    my  = ext.height() * 0.02
    bbox_comm = (f"{ext.xMinimum()-mx},{ext.yMinimum()-my},"
                 f"{ext.xMaximum()+mx},{ext.yMaximum()+my}")

    uri_comm = (f"url='{WFS_GPF}' "
                f"typename='LIMITES_ADMINISTRATIVES_EXPRESS.LATEST:commune' "
                f"version='1.1.0' srsname='EPSG:4326' restrictToRequestBBOX='1' "
                f"bbox='{bbox_comm},EPSG:4326'")
    lyr_comm = QgsVectorLayer(uri_comm, "communes_tmp", "WFS")
    if not lyr_comm.isValid():
        return None

    import processing
    try:
        comm_clipped = processing.run("native:clip", {
            'INPUT': lyr_comm, 'OVERLAY': lyr_territoire, 'OUTPUT': 'memory:'
        })['OUTPUT']
    except Exception:
        comm_clipped = lyr_comm

    codes_insee = sorted(set(
        str(f['code_insee'])
        for f in comm_clipped.getFeatures()
        if f['code_insee']
    ))

    lbl_status.setText(f"GPU — codes INSEE trouvés : {codes_insee} | {nom}")
    app.processEvents()

    if not codes_insee:
        lbl_status.setText(f"⚠ {nom} : aucun code INSEE trouvé dans le territoire")
        app.processEvents()
        return None

    lbl_status.setText(f"GPU — {len(codes_insee)} communes | {nom}...")
    app.processEvents()

    # Étape 2 — charger par lots via filtre OGC XML sur partition
    all_feats  = []
    ref_fields = None
    ref_crs    = None
    ref_geom   = None
    batch_size = 50

    for i in range(0, len(codes_insee), batch_size):
        batch = codes_insee[i:i+batch_size]

        # Construire le filtre OGC XML — OR sur les partitions
        if len(batch) == 1:
            filtre_partition = (
                f"<PropertyIsEqualTo>"
                f"<PropertyName>partition</PropertyName>"
                f"<Literal>DU_{batch[0]}</Literal>"
                f"</PropertyIsEqualTo>"
            )
        else:
            conditions = "".join(
                f"<PropertyIsEqualTo>"
                f"<PropertyName>partition</PropertyName>"
                f"<Literal>DU_{code}</Literal>"
                f"</PropertyIsEqualTo>"
                for code in batch
            )
            filtre_partition = f"<Or>{conditions}</Or>"

        filtre_ogc = f"<Filter>{filtre_partition}</Filter>"

        uri_batch = (f"url='{WFS_GPF}' typename='{typename}' version='1.1.0' "
                     f"srsname='EPSG:4326' filter='{filtre_ogc}'")
        lyr_batch = QgsVectorLayer(uri_batch, "gpu_batch", "WFS")

        if not lyr_batch.isValid():
            continue

        if ref_fields is None:
            ref_fields = lyr_batch.fields().toList()
            ref_crs    = lyr_batch.crs().authid()

            if lyr_batch.geometryType() == 0:
                ref_geom = "Point"
            elif lyr_batch.geometryType() == 1:
                ref_geom = "LineString"
            else:
                ref_geom = "Polygon"

        # Filtre attributaire si nécessaire (typepsc, stypepsc)
        if filtre_expr:
            expr    = QgsExpression(filtre_expr)
            request = QgsFeatureRequest(expr)
            feats   = list(lyr_batch.getFeatures(request))
        else:
            feats = list(lyr_batch.getFeatures())

        if not feats:
            continue

        all_feats.extend(feats)
        lbl_status.setText(
            f"GPU — lot {i//batch_size+1}/{(len(codes_insee)-1)//batch_size+1} "
            f"| {len(all_feats)} entités | {nom}")
        app.processEvents()

    if not all_feats or ref_fields is None:
        lbl_status.setText(
            f"ℹ {nom} : aucune donnée — commune(s) sans PLU numérique sur le GPU")
        app.processEvents()
        return None

    # Étape 3 — couche mémoire résultat
    lyr_result = QgsVectorLayer(f"{ref_geom}?crs={ref_crs}", nom, "memory")
    pr = lyr_result.dataProvider()
    pr.addAttributes(ref_fields)
    lyr_result.updateFields()
    pr.addFeatures(all_feats)
    lyr_result.updateExtents()
    return lyr_result
class ChargeurDonnees(QDialog):
    def __init__(self):
        super().__init__(iface.mainWindow())
        self.setWindowTitle("BioBlio")
        self.setMinimumWidth(700)
        self.resize(760, 820)
        self._territory_extent = None
        self._territory_lyr    = None
        self._territory_code   = None
        self._suggestions      = {}
        self._heavy_layer_rules = {
            "🧾 Parcelles cadastrales": ["Département", "Région", "France entière"],
            "🏠 Bâtiments cadastraux": ["Département", "Région", "France entière"],
            "🟦 Zones PLU / Cartes communales": ["Région", "France entière"],
            "🌲 Espaces Boisés Classés (EBC)": ["Région", "France entière"],
            "🌿 Éléments paysagers à préserver": ["Région", "France entière"],
            "🌳 Arbres / éléments naturels ponctuels à préserver": ["Région", "France entière"],
            "🏛 Monuments historiques — AC1": ["Région", "France entière"],
            "🏞 Sites classés / inscrits — AC2": ["Région", "France entière"],
            "🏘 Sites patrimoniaux remarquables — AC4": ["Région", "France entière"],
        }
        self._build_ui()

    def _build_ui(self):
        self._excel_export_path = ""

        # ── Palette Dark Pro ──────────────────────────────────────────────────
        BG       = "#1E1E2E"   # fond principal
        BG_HDR   = "#161625"   # en-tête et barre du bas
        BG_SEC   = "#252840"   # fond des corps d'accordéon
        BG_ACC   = "#1E2235"   # fond alterné (corps)
        BG_INP   = "#13141F"   # champs de saisie
        BG_HOV   = "#2E3148"   # survol items
        TXT      = "#E2E8F0"   # texte principal
        TXT_DIM  = "#94A3B8"   # texte secondaire
        ACCENT   = "#10B981"   # vert émeraude
        ACCENT_H = "#059669"
        ACCENT_P = "#047857"
        BORDER   = "#383B52"   # bordures

        self.setStyleSheet(f"QDialog {{ background:{BG}; }}")

        FIELD = (
            f"QLineEdit {{ background:{BG_INP}; color:{TXT}; "
            f"border:1px solid {BORDER}; border-radius:5px; padding:5px 9px; }}"
            f"QLineEdit:focus {{ border-color:{ACCENT}; }}"
            f"QLineEdit::placeholder {{ color:{TXT_DIM}; }}"
        )
        COMBO = (
            f"QComboBox {{ background:{BG_INP}; color:{TXT}; "
            f"border:1px solid {BORDER}; border-radius:5px; padding:5px 9px; }}"
            f"QComboBox:hover {{ border-color:{ACCENT}; }}"
            f"QComboBox QAbstractItemView {{ background:{BG_SEC}; color:{TXT}; "
            f"selection-background-color:{ACCENT}; border:1px solid {BORDER}; }}"
        )
        CHK = (
            f"QCheckBox {{ color:{TXT}; padding:3px 6px; border-radius:4px; spacing:6px; }}"
            f"QCheckBox:hover {{ background:{BG_HOV}; }}"
            f"QCheckBox::indicator {{ width:14px; height:14px; border-radius:3px; "
            f"border:1px solid {BORDER}; background:{BG_INP}; }}"
            f"QCheckBox::indicator:checked {{ background:{ACCENT}; border-color:{ACCENT}; }}"
            f"QCheckBox::indicator:disabled {{ background:{BORDER}; border-color:{BORDER}; }}"
        )
        BTN_SM = (
            f"QPushButton {{ background:{BG_HOV}; color:{TXT_DIM}; "
            f"border:1px solid {BORDER}; border-radius:4px; padding:3px 10px; font-size:11px; }}"
            f"QPushButton:hover {{ border-color:{ACCENT}; color:{ACCENT}; }}"
        )

        def make_accordion(title, expanded=True):
            """Retourne (outer_widget, body_layout) pour une section accordéon."""
            outer = QWidget()
            outer.setObjectName("acc_outer")
            outer.setStyleSheet("QWidget#acc_outer { background:transparent; }")
            vlay = QVBoxLayout(outer)
            vlay.setSpacing(0)
            vlay.setContentsMargins(0, 0, 0, 4)

            btn = QPushButton()
            btn.setCursor(Qt.PointingHandCursor)

            def _set_style(is_open):
                r_bottom = "0 0" if is_open else "6px 6px"
                btn.setStyleSheet(
                    f"QPushButton {{ background:{BG_SEC}; color:{TXT}; border:none; "
                    f"border-radius:6px 6px {r_bottom}; padding:10px 14px; "
                    f"text-align:left; font-weight:bold; font-size:12px; }}"
                    f"QPushButton:hover {{ background:{BG_HOV}; }}"
                )
                icon = "▼  " if is_open else "▶  "
                btn.setText(icon + title)

            _set_style(expanded)

            body = QWidget()
            body.setObjectName("acc_body")
            body.setStyleSheet(
                f"QWidget#acc_body {{ background:{BG_ACC}; "
                f"border-radius:0 0 6px 6px; }}"
            )
            body_layout = QVBoxLayout(body)
            body_layout.setContentsMargins(12, 10, 12, 12)
            body_layout.setSpacing(8)
            body.setVisible(expanded)

            def toggle():
                is_open = not body.isVisible()
                body.setVisible(is_open)
                _set_style(is_open)

            btn.clicked.connect(toggle)
            vlay.addWidget(btn)
            vlay.addWidget(body)
            return outer, body_layout

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # ── En-tête ───────────────────────────────────────────────────────────
        header = QWidget()
        header.setFixedHeight(52)
        header.setStyleSheet(f"background:{BG_HDR};")
        lay_h = QHBoxLayout(header)
        lay_h.setContentsMargins(16, 0, 16, 0)
        import os as _os
        lbl_icon = QLabel()
        lbl_icon.setStyleSheet("background:transparent;")
        _icon_path = _os.path.join(_os.path.dirname(__file__), "icon.png")
        if _os.path.exists(_icon_path):
            _pix = QPixmap(_icon_path).scaled(
                36, 36, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            lbl_icon.setPixmap(_pix)
        lay_h.addWidget(lbl_icon)
        lbl_title = QLabel(" BioBlio")
        lbl_title.setFont(QFont("Arial", 13, QFont.Bold))
        lbl_title.setStyleSheet(f"color:{TXT}; background:transparent;")
        lay_h.addWidget(lbl_title)
        lbl_sub = QLabel("  —  La bibliographie environnementale en un clic")
        lbl_sub.setStyleSheet(f"color:{TXT_DIM}; font-size:10px; background:transparent;")
        lay_h.addWidget(lbl_sub)
        lay_h.addStretch()
        root_layout.addWidget(header)

        accent_line = QFrame()
        accent_line.setFixedHeight(2)
        accent_line.setStyleSheet(f"background:{ACCENT};")
        root_layout.addWidget(accent_line)

        # ── Zone scrollable ───────────────────────────────────────────────────
        global_scroll = QScrollArea()
        global_scroll.setWidgetResizable(True)
        global_scroll.setFrameShape(QFrame.NoFrame)
        global_scroll.setStyleSheet(
            f"QScrollArea {{ background:{BG}; border:none; }}"
            f"QScrollBar:vertical {{ background:{BG}; width:7px; border:none; }}"
            f"QScrollBar::handle:vertical {{ background:{BORDER}; border-radius:3px; min-height:20px; }}"
            f"QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0; }}"
        )
        scroll_container = QWidget()
        scroll_container.setStyleSheet(f"background:{BG};")
        main = QVBoxLayout(scroll_container)
        main.setSpacing(0)
        main.setContentsMargins(10, 10, 10, 6)

        # ── Accordéon 1 : Territoire ──────────────────────────────────────────
        sec1, lay_scale = make_accordion("📍  Territoire", expanded=True)

        self.combo_scale = QComboBox()
        self.combo_scale.addItems(["Commune", "Département", "Région", "France entière", "Zone personnalisée"])
        self.combo_scale.setStyleSheet(COMBO)
        self.combo_scale.currentIndexChanged.connect(self._on_scale_change)
        lay_scale.addWidget(self.combo_scale)

        lay_search = QHBoxLayout()
        lay_search.setSpacing(6)
        self.lbl_search = QLabel("Commune :")
        self.lbl_search.setFixedWidth(88)
        self.lbl_search.setStyleSheet(f"color:{TXT_DIM};")
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("ex: Tours (37)...")
        self.txt_search.setStyleSheet(FIELD)
        self._completer_model = QStringListModel()
        self._completer = QCompleter()
        self._completer.setModel(self._completer_model)
        self._completer.setCaseSensitivity(Qt.CaseInsensitive)
        self._completer.setFilterMode(Qt.MatchContains)
        self.txt_search.setCompleter(self._completer)
        self.txt_search.textChanged.connect(self._on_text_changed)

        self.btn_search = QPushButton("🔍 Rechercher")
        self.btn_search.setStyleSheet(
            f"QPushButton {{ background:{ACCENT}; color:white; border:none; "
            f"border-radius:5px; padding:6px 14px; font-weight:bold; }}"
            f"QPushButton:hover {{ background:{ACCENT_H}; }}"
            f"QPushButton:pressed {{ background:{ACCENT_P}; }}"
            f"QPushButton:disabled {{ background:#2A4A40; color:#567A6A; }}")
        self.btn_search.clicked.connect(self._search_territory)

        self.btn_browse = QPushButton("📂 Fichier")
        self.btn_browse.setStyleSheet(
            f"QPushButton {{ background:#92400E; color:white; border:none; "
            f"border-radius:5px; padding:6px 10px; }}"
            f"QPushButton:hover {{ background:#B45309; }}")
        self.btn_browse.clicked.connect(self._browse_zone)
        self.btn_browse.setVisible(False)

        self.btn_from_layer = QPushButton("🗂 Projet")
        self.btn_from_layer.setStyleSheet(
            f"QPushButton {{ background:#1E3A5F; color:white; border:none; "
            f"border-radius:5px; padding:6px 10px; }}"
            f"QPushButton:hover {{ background:#2563EB; }}")
        self.btn_from_layer.clicked.connect(self._use_project_layer)
        self.btn_from_layer.setVisible(False)

        lay_search.addWidget(self.lbl_search)
        lay_search.addWidget(self.txt_search)
        lay_search.addWidget(self.btn_search)
        lay_search.addWidget(self.btn_browse)
        lay_search.addWidget(self.btn_from_layer)
        lay_scale.addLayout(lay_search)

        lay_layer_pick = QHBoxLayout()
        self.combo_layers = QComboBox()
        self.combo_layers.setMinimumWidth(300)
        self.combo_layers.setStyleSheet(COMBO)
        self.combo_layers.setVisible(False)
        lay_layer_pick.addWidget(self.combo_layers)
        lay_layer_pick.addStretch()
        lay_scale.addLayout(lay_layer_pick)

        self.chk_buffer = QCheckBox("Appliquer un buffer autour du territoire")
        self.chk_buffer.setChecked(False)
        self.chk_buffer.setStyleSheet(CHK)
        self.chk_buffer.stateChanged.connect(self._on_buffer_toggle)
        lay_scale.addWidget(self.chk_buffer)

        self.buffer_widget = QWidget()
        self.buffer_widget.setStyleSheet("background:transparent;")
        lay_buffer = QHBoxLayout(self.buffer_widget)
        lay_buffer.setContentsMargins(22, 0, 0, 0)
        lay_buffer.setSpacing(8)
        lbl_buf = QLabel("Taille :")
        lbl_buf.setStyleSheet(f"color:{TXT_DIM};")
        lay_buffer.addWidget(lbl_buf)
        self.txt_buffer_size = QLineEdit()
        self.txt_buffer_size.setPlaceholderText("ex: 10")
        self.txt_buffer_size.setFixedWidth(65)
        self.txt_buffer_size.setStyleSheet(FIELD)
        lay_buffer.addWidget(self.txt_buffer_size)
        self.combo_buffer_unit = QComboBox()
        self.combo_buffer_unit.addItems(["Kilomètres", "Mètres"])
        self.combo_buffer_unit.setFixedWidth(110)
        self.combo_buffer_unit.setStyleSheet(COMBO)
        lay_buffer.addWidget(self.combo_buffer_unit)
        lay_buffer.addStretch()
        self.buffer_widget.setVisible(False)
        lay_scale.addWidget(self.buffer_widget)

        self.chk_intersect_only = QCheckBox(
            "Conserver les entités entières (sans découpage exact)")
        self.chk_intersect_only.setChecked(False)
        self.chk_intersect_only.setStyleSheet(CHK)
        self.chk_intersect_only.setToolTip(
            "Coché : entités qui touchent le territoire chargées entières\n"
            "Décoché : entités découpées exactement sur le contour")
        lay_scale.addWidget(self.chk_intersect_only)

        self.lbl_result = QLabel("")
        self.lbl_result.setStyleSheet(f"color:{ACCENT}; font-style:italic; padding:2px;")
        lay_scale.addWidget(self.lbl_result)
        main.addWidget(sec1)

        # ── Accordéon 2 : Couches à charger ──────────────────────────────────
        sec2, lay_data = make_accordion("📋  Couches à charger", expanded=True)

        lay_btns = QHBoxLayout()
        btn_all  = QPushButton("✔  Tout cocher")
        btn_none = QPushButton("✗  Tout décocher")
        btn_all.setStyleSheet(BTN_SM)
        btn_none.setStyleSheet(BTN_SM)
        btn_all.clicked.connect(lambda: self._check_all(True))
        btn_none.clicked.connect(lambda: self._check_all(False))
        lay_btns.addWidget(btn_all)
        lay_btns.addWidget(btn_none)
        lay_btns.addStretch()
        lay_data.addLayout(lay_btns)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMinimumHeight(240)
        scroll.setMaximumHeight(420)
        scroll.setStyleSheet(
            f"QScrollArea {{ border:1px solid {BORDER}; border-radius:4px; background:{BG_INP}; }}"
            f"QScrollBar:vertical {{ background:{BG_INP}; width:6px; border:none; }}"
            f"QScrollBar::handle:vertical {{ background:{BORDER}; border-radius:3px; }}"
            f"QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0; }}"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{BG_INP};")
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(1)
        scroll_layout.setContentsMargins(6, 6, 6, 6)

        self.checkboxes = {}
        current_cat = None
        for cat, nom, *_ in CATALOGUE:
            if cat != current_cat:
                lbl_cat = QLabel(cat)
                lbl_cat.setFont(QFont("Arial", 8, QFont.Bold))
                lbl_cat.setStyleSheet(
                    f"color:{ACCENT}; padding:6px 4px 2px 4px; "
                    f"border-bottom:1px solid {BORDER};")
                scroll_layout.addWidget(lbl_cat)
                current_cat = cat
            cb = QCheckBox(f"  {nom}")
            cb.setChecked(False)
            cb.setStyleSheet(CHK)
            self.checkboxes[nom] = cb
            scroll_layout.addWidget(cb)

        self._update_layer_availability()
        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        lay_data.addWidget(scroll)
        main.addWidget(sec2)

        # ── Accordéon 3 : Options ─────────────────────────────────────────────
        sec3, lay_opt_body = make_accordion("⚙  Options", expanded=False)

        lay_opt = QHBoxLayout()
        lay_opt.setSpacing(20)
        self.chk_clip  = QCheckBox("✂  Découper sur le territoire")
        self.chk_style = QCheckBox("🎨  Appliquer la symbologie")
        self.chk_group = QCheckBox("📁  Grouper les couches")
        self.chk_clip.setChecked(True)
        self.chk_style.setChecked(True)
        self.chk_group.setChecked(True)
        self.chk_clip.setToolTip("Découper les couches exactement sur le contour du territoire")
        self.chk_style.setToolTip("Appliquer automatiquement la symbologie définie")
        self.chk_group.setToolTip("Regrouper les couches chargées dans un groupe nommé")
        for chk in (self.chk_clip, self.chk_style, self.chk_group):
            chk.setStyleSheet(CHK)
            lay_opt.addWidget(chk)
        lay_opt.addStretch()
        lay_opt_body.addLayout(lay_opt)
        main.addWidget(sec3)

        # ── Accordéon 4 : GBIF ────────────────────────────────────────────────
        sec4, lay_gbif_body = make_accordion("🦋  Observations GBIF", expanded=False)

        self.chk_gbif = QCheckBox("Charger les observations GBIF sur le territoire")
        self.chk_gbif.setChecked(False)
        self.chk_gbif.setStyleSheet(CHK)
        self.chk_gbif.stateChanged.connect(self._on_gbif_toggle)
        lay_gbif_body.addWidget(self.chk_gbif)

        self.gbif_widget = QWidget()
        self.gbif_widget.setStyleSheet("background:transparent;")
        lay_gbif_inner = QVBoxLayout(self.gbif_widget)
        lay_gbif_inner.setContentsMargins(0, 4, 0, 0)
        lay_gbif_inner.setSpacing(8)

        lbl_taxon = QLabel("Groupes taxonomiques (multi-sélection) :")
        lbl_taxon.setStyleSheet(f"color:{TXT_DIM}; font-size:11px;")
        lay_gbif_inner.addWidget(lbl_taxon)

        taxon_scroll = QScrollArea()
        taxon_scroll.setWidgetResizable(True)
        taxon_scroll.setFixedHeight(160)
        taxon_scroll.setStyleSheet(
            f"QScrollArea {{ border:1px solid {BORDER}; border-radius:4px; background:{BG_INP}; }}"
            f"QScrollBar:vertical {{ background:{BG_INP}; width:6px; border:none; }}"
            f"QScrollBar::handle:vertical {{ background:{BORDER}; border-radius:3px; }}"
        )
        taxon_widget = QWidget()
        taxon_widget.setStyleSheet(f"background:{BG_INP};")
        taxon_layout = QVBoxLayout(taxon_widget)
        taxon_layout.setSpacing(1)
        taxon_layout.setContentsMargins(4, 2, 4, 2)

        self.gbif_checkboxes = {}
        for label, key in GBIF_GROUPES:
            cb = QCheckBox(label)
            cb.setChecked(False)
            cb.setStyleSheet(CHK)
            if label == "Tous groupes":
                cb.setChecked(True)
                cb.stateChanged.connect(self._on_tous_groupes_toggle)
            taxon_layout.addWidget(cb)
            self.gbif_checkboxes[label] = (cb, key)

        taxon_widget.setLayout(taxon_layout)
        taxon_scroll.setWidget(taxon_widget)
        lay_gbif_inner.addWidget(taxon_scroll)

        sep_gbif = QFrame()
        sep_gbif.setFrameShape(QFrame.HLine)
        sep_gbif.setStyleSheet(f"color:{BORDER};")
        lay_gbif_inner.addWidget(sep_gbif)

        lay_annees = QHBoxLayout()
        lbl_ann = QLabel("Années :")
        lbl_ann.setStyleSheet(f"color:{TXT_DIM};")
        lay_annees.addWidget(lbl_ann)
        self.txt_gbif_annee_min = QLineEdit()
        self.txt_gbif_annee_min.setPlaceholderText("ex: 2000")
        self.txt_gbif_annee_min.setFixedWidth(70)
        self.txt_gbif_annee_min.setStyleSheet(FIELD)
        lay_annees.addWidget(self.txt_gbif_annee_min)
        lbl_to = QLabel(" → ")
        lbl_to.setStyleSheet(f"color:{TXT_DIM};")
        lay_annees.addWidget(lbl_to)
        self.txt_gbif_annee_max = QLineEdit()
        self.txt_gbif_annee_max.setPlaceholderText("ex: 2024")
        self.txt_gbif_annee_max.setFixedWidth(70)
        self.txt_gbif_annee_max.setStyleSheet(FIELD)
        lay_annees.addWidget(self.txt_gbif_annee_max)
        lay_annees.addStretch()
        lay_gbif_inner.addLayout(lay_annees)

        lay_mois = QHBoxLayout()
        lbl_mois = QLabel("Mois :")
        lbl_mois.setStyleSheet(f"color:{TXT_DIM};")
        lay_mois.addWidget(lbl_mois)
        MOIS = ["—", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        self.combo_gbif_mois_min = QComboBox()
        self.combo_gbif_mois_min.addItems(MOIS)
        self.combo_gbif_mois_min.setFixedWidth(110)
        self.combo_gbif_mois_min.setStyleSheet(COMBO)
        lay_mois.addWidget(self.combo_gbif_mois_min)
        lbl_to2 = QLabel(" → ")
        lbl_to2.setStyleSheet(f"color:{TXT_DIM};")
        lay_mois.addWidget(lbl_to2)
        self.combo_gbif_mois_max = QComboBox()
        self.combo_gbif_mois_max.addItems(MOIS)
        self.combo_gbif_mois_max.setFixedWidth(110)
        self.combo_gbif_mois_max.setStyleSheet(COMBO)
        lay_mois.addWidget(self.combo_gbif_mois_max)
        lay_mois.addStretch()
        lay_gbif_inner.addLayout(lay_mois)

        lay_limit = QHBoxLayout()
        lbl_lim = QLabel("Nombre max d'occurrences :")
        lbl_lim.setStyleSheet(f"color:{TXT_DIM};")
        lay_limit.addWidget(lbl_lim)
        self.combo_gbif_limit = QComboBox()
        self.combo_gbif_limit.addItems(["300", "1 000", "3 000", "10 000", "50 000", "100 000"])
        self.combo_gbif_limit.setCurrentIndex(1)
        self.combo_gbif_limit.setStyleSheet(COMBO)
        lay_limit.addWidget(self.combo_gbif_limit)
        lay_limit.addStretch()
        lay_gbif_inner.addLayout(lay_limit)

        self.gbif_widget.setVisible(False)
        lay_gbif_body.addWidget(self.gbif_widget)
        main.addWidget(sec4)

        main.addStretch()
        scroll_container.setLayout(main)
        global_scroll.setWidget(scroll_container)
        root_layout.addWidget(global_scroll)

        # ── Barre du bas (fixe, hors scroll) ─────────────────────────────────
        bottom_bar = QWidget()
        bottom_bar.setStyleSheet(
            f"background:{BG_HDR}; border-top:1px solid {BORDER};")
        lay_bottom = QVBoxLayout(bottom_bar)
        lay_bottom.setContentsMargins(10, 8, 10, 10)
        lay_bottom.setSpacing(5)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setFixedHeight(3)
        self.progress.setTextVisible(False)
        self.progress.setStyleSheet(
            f"QProgressBar {{ border:none; background:{BG}; }}"
            f"QProgressBar::chunk {{ background:{ACCENT}; }}"
        )
        lay_bottom.addWidget(self.progress)

        self.lbl_status = QLabel("")
        self.lbl_status.setStyleSheet(
            f"color:{TXT_DIM}; font-size:11px; font-style:italic;")
        lay_bottom.addWidget(self.lbl_status)

        lay_buttons = QHBoxLayout()
        lay_buttons.setSpacing(8)

        def _btn(label, bg, bg_h, bg_p):
            b = QPushButton(label)
            b.setMinimumHeight(40)
            b.setStyleSheet(
                f"QPushButton {{ background:{bg}; color:white; border:none; "
                f"border-radius:6px; padding:8px 6px; font-size:12px; font-weight:bold; }}"
                f"QPushButton:hover {{ background:{bg_h}; }}"
                f"QPushButton:pressed {{ background:{bg_p}; }}"
                f"QPushButton:disabled {{ background:{BG_HOV}; color:{BORDER}; }}"
            )
            return b

        self.btn_load = _btn("🚀  Charger", ACCENT, ACCENT_H, ACCENT_P)
        self.btn_load.clicked.connect(self._load_data)
        lay_buttons.addWidget(self.btn_load)

        self.btn_save_gpkg = _btn("💾  GeoPackage", "#2563EB", "#1D4ED8", "#1E40AF")
        self.btn_save_gpkg.clicked.connect(self._export_to_gpkg)
        lay_buttons.addWidget(self.btn_save_gpkg)

        self.btn_export_excel = _btn("📊  Excel", "#D97706", "#B45309", "#92400E")
        self.btn_export_excel.clicked.connect(self._export_excel_from_project)
        lay_buttons.addWidget(self.btn_export_excel)

        lay_bottom.addLayout(lay_buttons)
        root_layout.addWidget(bottom_bar)

    def _msgbox(self, level, title, text):
        """Affiche un QMessageBox stylé dark — texte blanc garanti."""
        icons = {
            "info":     QMessageBox.Information,
            "warning":  QMessageBox.Warning,
            "critical": QMessageBox.Critical,
        }
        msg = QMessageBox(self)
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setIcon(icons.get(level, QMessageBox.Information))
        # Stylesheet — QMessageBox QLabel nécessaire sur Windows
        msg.setStyleSheet(
            "QMessageBox { background-color:#1E2235; }"
            "QMessageBox QLabel { color:#E2E8F0; font-size:12px; }"
            "QMessageBox QPushButton { background:#10B981; color:white; border:none; "
            "border-radius:4px; padding:5px 20px; min-width:60px; font-weight:bold; }"
            "QMessageBox QPushButton:hover { background:#059669; }"
        )
        # Palette en fallback pour les textes que Qt gère hors stylesheet
        pal = msg.palette()
        pal.setColor(QPalette.WindowText, QColor("#E2E8F0"))
        pal.setColor(QPalette.Window,     QColor("#1E2235"))
        pal.setColor(QPalette.ButtonText, QColor("white"))
        pal.setColor(QPalette.Button,     QColor("#10B981"))
        msg.setPalette(pal)
        msg.exec_()

    def _browse_zone(self):
        """Charge un fichier vecteur comme zone personnalisée."""
        from qgis.PyQt.QtWidgets import QFileDialog
        from qgis.core import QgsVectorLayer

        # Formats acceptés : GeoPackage, GeoJSON, Shapefile, KML
        filtres = (
            "Fichiers vecteur (*.gpkg *.geojson *.json *.shp *.kml *.kmz);;"
            "GeoPackage (*.gpkg);;"
            "GeoJSON (*.geojson *.json);;"
            "Shapefile (*.shp);;"
            "KML/KMZ (*.kml *.kmz);;"
            "Tous les fichiers (*)"
        )
        chemin, _ = QFileDialog.getOpenFileName(
            self, "Choisir une zone personnalisée", "", filtres)

        if not chemin:
            return

        lyr = QgsVectorLayer(chemin, "zone_perso", "ogr")

        if not lyr.isValid() or lyr.featureCount() == 0:
            self.lbl_result.setText("❌ Fichier invalide ou vide.")
            self.lbl_result.setStyleSheet("color:red; font-style:italic;")
            return

        if lyr.geometryType() != 2:  # 2 = polygone
            self.lbl_result.setText("❌ Le fichier doit contenir des polygones.")
            self.lbl_result.setStyleSheet("color:red; font-style:italic;")
            return

        # Reprojeter en WGS84 si nécessaire pour cohérence avec les WFS
        from qgis.core import QgsCoordinateReferenceSystem
        if lyr.crs().authid() != "EPSG:4326":
            import processing
            res = processing.run("native:reprojectlayer", {
                'INPUT': lyr, 'TARGET_CRS': 'EPSG:4326', 'OUTPUT': 'memory:'
            })
            lyr = res['OUTPUT']

        # Utiliser l'extent de l'union de tous les polygones
        self._territory_extent = lyr.extent()
        self._territory_lyr    = lyr

        import os
        nom_fichier = os.path.splitext(os.path.basename(chemin))[0]
        self.txt_search.setText(nom_fichier)  # afficher le nom du fichier
        self.txt_search.setEnabled(False)

        n = lyr.featureCount()
        self.lbl_result.setText(
            f"✅ Zone chargée : {nom_fichier} ({n} polygone{'s' if n > 1 else ''})")
        self.lbl_result.setStyleSheet("color:#2A7A2A; font-style:italic;")

    def _refresh_layer_combo(self):
        """Remplit le combo avec les couches polygones du projet."""
        from qgis.core import QgsProject
        self.combo_layers.clear()
        self._layer_map = {}  # nom affiché → objet QgsVectorLayer
        for lid, lyr in QgsProject.instance().mapLayers().items():
            if hasattr(lyr, 'geometryType') and lyr.geometryType() == 2:
                label = f"{lyr.name()} ({lyr.featureCount()} entités)"
                self.combo_layers.addItem(label)
                self._layer_map[label] = lyr
        if not self._layer_map:
            self.combo_layers.addItem("Aucune couche polygone dans le projet")

    def _use_project_layer(self):
        """Utilise la couche sélectionnée dans le combo comme zone."""
        import processing
        from qgis.core import QgsCoordinateReferenceSystem

        label = self.combo_layers.currentText()
        lyr = self._layer_map.get(label)
        if not lyr:
            self.lbl_result.setText("❌ Aucune couche sélectionnée.")
            self.lbl_result.setStyleSheet("color:red; font-style:italic;")
            return

        # Reprojeter en WGS84 si nécessaire
        if lyr.crs().authid() != "EPSG:4326":
            res = processing.run("native:reprojectlayer", {
                'INPUT': lyr, 'TARGET_CRS': 'EPSG:4326', 'OUTPUT': 'memory:'
            })
            lyr_wgs = res['OUTPUT']
        else:
            lyr_wgs = lyr

        self._territory_extent = lyr_wgs.extent()
        self._territory_lyr    = lyr_wgs

        self.txt_search.setText(lyr.name())
        self.lbl_result.setText(
            f"✅ Zone : {lyr.name()} ({lyr.featureCount()} polygone(s))")
        self.lbl_result.setStyleSheet("color:#2A7A2A; font-style:italic;")

    def _on_buffer_toggle(self, state):
        """Affiche ou masque les options de buffer."""
        self.buffer_widget.setVisible(state == Qt.Checked)

    def _update_layer_availability(self):
        """Désactive les couches trop lourdes aux grandes échelles."""
        if not hasattr(self, "checkboxes"):
            return

        scale = self.combo_scale.currentText()

        for nom, cb in self.checkboxes.items():
            disabled_scales = self._heavy_layer_rules.get(nom, [])
            is_disabled = scale in disabled_scales

            cb.setEnabled(not is_disabled)

            if is_disabled:
                cb.setChecked(False)
                cb.setToolTip(
                    "Couche désactivée à cette échelle : elle est trop lourde "
                    "à charger sur un territoire aussi vaste."
                )
            else:
                cb.setToolTip("")

    def _on_text_changed(self, text):
        """Autocomplétion dynamique via API Géo gouvernement."""
        import requests
        self._suggestions = {}

        if len(text) < 3:
            self._completer_model.setStringList([])
            return
    
        scale = self.combo_scale.currentText()

        try:
            if scale == "Commune":
                r = requests.get(
                    "https://geo.api.gouv.fr/communes",
                    params={"nom": text, "fields": "nom,code,codeDepartement",
                            "boost": "population", "limit": 12},
                    timeout=5)
                if r.status_code == 200:
                    data = r.json()
                    suggestions = []
                    for item in data:
                        # Affiche : "Tours (37)"
                        label = f"{item['nom']} ({item.get('codeDepartement', '')})"
                        self._suggestions[label] = item['code']
                        suggestions.append(label)
                    self._completer_model.setStringList(suggestions)

            elif scale == "Département":
                r = requests.get(
                    "https://geo.api.gouv.fr/departements",
                    params={"nom": text, "fields": "nom,code", "limit": 15},
                    timeout=5)
                if r.status_code == 200:
                    data = r.json()
                    suggestions = []
                    for item in data:
                        label = f"{item['nom']} ({item['code']})"
                        self._suggestions[label] = item['code']
                        suggestions.append(label)
                    self._completer_model.setStringList(suggestions)

            elif scale == "Région":
                r = requests.get(
                    "https://geo.api.gouv.fr/regions",
                    params={"nom": text, "fields": "nom,code", "limit": 18},
                    timeout=5)
                if r.status_code == 200:
                    data = r.json()
                    suggestions = []
                    for item in data:
                        # Stocker le nom seul comme clé (sans parenthèse)
                        # car le WFS cherche sur nom_officiel
                        label = f"{item['nom']} ({item['code']})"
                        # Stocker le nom propre pour le fallback nom_officiel
                        self._suggestions[label] = item['code']
                        self._suggestions[item['nom']] = item['code']
                        suggestions.append(label)
                    self._completer_model.setStringList(suggestions)
        except Exception:
            pass  # pas de réseau ou timeout → on ignore silencieusement
        
    def _on_scale_change(self, idx):
        labels = ["Nom de la commune :", "Nom ou N° du département :",
                  "Nom de la région :", "", "Fichier de zone :"]
        placeholders = ["ex: Tours (37)...", "ex: Indre-et-Loire ou 37",
                        "ex: Centre (tapez 3 lettres)...", "",
                        "Charger un fichier de zone..."]
        self.lbl_search.setText(labels[idx])
        self.txt_search.setPlaceholderText(placeholders[idx])

        # Réinitialiser
        self._territory_extent = None
        self._territory_lyr    = None
        self._territory_code   = None
        self._suggestions      = {}
        self.txt_search.clear()
        self._completer_model.setStringList([])
        self.lbl_result.setText("")
        self._update_layer_availability()

        if idx == 3:  # France entière
            self.txt_search.setEnabled(False)
            self.btn_search.setEnabled(False)
            self.btn_browse.setVisible(False)
            self.btn_from_layer.setVisible(False)   # ← ajouter
            self.combo_layers.setVisible(False)      # ← ajouter
            self.lbl_result.setText("✅ Chargement sur la France entière")
            self.lbl_result.setStyleSheet("color:#2A7A2A; font-style:italic;")
        elif idx == 4:  # Zone personnalisée
            self.txt_search.setEnabled(False)
            self.btn_search.setEnabled(False)
            self.btn_browse.setVisible(True)
            self.btn_from_layer.setVisible(True)
            self._refresh_layer_combo()
            self.combo_layers.setVisible(True)
        else:
            self.txt_search.setEnabled(True)
            self.btn_search.setEnabled(True)
            self.btn_browse.setVisible(False)
            self.btn_from_layer.setVisible(False)   # ← ajouter
            self.combo_layers.setVisible(False)      # ← ajouter

    def _search_territory(self):
        import html as _html
        scale = self.combo_scale.currentText()
        terme = self.txt_search.text().strip()

        if not terme:
            self._msgbox("warning", "Attention", "Saisissez un nom de territoire.")
            return

        # Vérifier que le terme correspond à une suggestion connue
        # (pour éviter une recherche WFS sur un texte arbitraire qui plante)
        code = self._suggestions.get(terme, None)
        if scale in ("Commune", "Région") and not code:
            self._msgbox("warning", "Territoire introuvable",
                f"Aucune {scale.lower()} ne correspond à « {terme} ».\n"
                "Tapez au moins 3 caractères et sélectionnez une suggestion dans la liste.")
            self.lbl_result.setText("")
            return
        if scale == "Département" and not code and not terme.isdigit():
            self._msgbox("warning", "Territoire introuvable",
                f"Aucun département ne correspond à « {terme} ».\n"
                "Tapez le nom ou le numéro et sélectionnez une suggestion dans la liste.")
            self.lbl_result.setText("")
            return

        self.lbl_result.setText("Recherche en cours...")
        self.lbl_result.setStyleSheet("color:#888; font-style:italic;")
        QApplication.processEvents()

        try:

            if scale == "Commune":
                typename = "LIMITES_ADMINISTRATIVES_EXPRESS.LATEST:commune"
                if code:
                    filtre = (f"<Filter><PropertyIsEqualTo>"
                              f"<PropertyName>code_insee</PropertyName>"
                              f"<Literal>{code}</Literal>"
                              f"</PropertyIsEqualTo></Filter>")
                else:
                    nom_seul = _html.escape(terme.split("(")[0].strip())
                    filtre = (f"<Filter><PropertyIsLike wildCard='*' "
                              f"singleChar='.' escape='!'>"
                              f"<PropertyName>nom_officiel</PropertyName>"
                              f"<Literal>{nom_seul}</Literal>"
                              f"</PropertyIsLike></Filter>")

            elif scale == "Département":
                typename = "LIMITES_ADMINISTRATIVES_EXPRESS.LATEST:departement"
                if code:
                    filtre = (f"<Filter><PropertyIsEqualTo>"
                              f"<PropertyName>code_insee</PropertyName>"
                              f"<Literal>{code}</Literal>"
                              f"</PropertyIsEqualTo></Filter>")
                elif terme.isdigit():
                    filtre = (f"<Filter><PropertyIsEqualTo>"
                              f"<PropertyName>code_insee</PropertyName>"
                              f"<Literal>{terme.zfill(2)}</Literal>"
                              f"</PropertyIsEqualTo></Filter>")
                else:
                    nom_seul = _html.escape(terme.split("(")[0].strip())
                    filtre = (f"<Filter><PropertyIsLike wildCard='*' "
                              f"singleChar='.' escape='!'>"
                              f"<PropertyName>nom_officiel</PropertyName>"
                              f"<Literal>*{nom_seul}*</Literal>"
                              f"</PropertyIsLike></Filter>")

            else:  # Région
                typename = "LIMITES_ADMINISTRATIVES_EXPRESS.LATEST:region"
                if code:
                    filtre = (f"<Filter><PropertyIsEqualTo>"
                              f"<PropertyName>code_insee</PropertyName>"
                              f"<Literal>{code}</Literal>"
                              f"</PropertyIsEqualTo></Filter>")
                else:
                    nom_seul = _html.escape(terme.split("(")[0].strip())
                    filtre = (f"<Filter><PropertyIsLike wildCard='*' "
                              f"singleChar='.' escape='!'>"
                              f"<PropertyName>nom_officiel</PropertyName>"
                              f"<Literal>*{nom_seul}*</Literal>"
                              f"</PropertyIsLike></Filter>")

            uri = (f"url='{WFS_GPF}' typename='{typename}' version='auto' "
                   f"srsname='EPSG:4326' filter='{filtre}'")
            lyr = QgsVectorLayer(uri, "search", "WFS")

            if not lyr.isValid() or lyr.featureCount() == 0:
                self.lbl_result.setText("❌ Territoire non trouvé.")
                self.lbl_result.setStyleSheet("color:red; font-style:italic;")
                self._territory_extent = None
                return

            nom_trouve = terme  # fallback si le champ est absent
            for feat in lyr.getFeatures():
                geom = feat.geometry()
                if geom and not geom.isEmpty():
                    self._territory_extent = geom.boundingBox()
                try:
                    nom_trouve = feat['nom_officiel'] or terme
                except Exception:
                    pass
                try:
                    code_feat = feat['code_insee']
                    self._territory_code = str(code_feat) if code_feat else code
                except Exception:
                    self._territory_code = code
                break

            if not self._territory_extent:
                self.lbl_result.setText("❌ Territoire trouvé mais géométrie invalide.")
                self.lbl_result.setStyleSheet("color:red; font-style:italic;")
                return

            # Convertir la couche WFS en mémoire pour éviter
            # le problème d'extent mondiale lors des clips
            lyr_copy = QgsVectorLayer(
                f"MultiPolygon?crs={lyr.crs().authid()}", "territoire", "memory")
            pr_copy = lyr_copy.dataProvider()
            pr_copy.addAttributes(lyr.fields().toList())
            lyr_copy.updateFields()
            pr_copy.addFeatures(list(lyr.getFeatures()))
            lyr_copy.updateExtents()
            self._territory_lyr = lyr_copy

            self.lbl_result.setText(f"✅ {nom_trouve}")
            self.lbl_result.setStyleSheet("color:#2A7A2A; font-style:italic;")

        except Exception:
            self.lbl_result.setText("❌ Erreur lors de la recherche.")
            self.lbl_result.setStyleSheet("color:red; font-style:italic;")
            self._territory_extent = None
            self._territory_lyr    = None
            self._territory_code   = None

    def _on_tous_groupes_toggle(self, state):
        """Si 'Tous groupes' est coché, décoche les autres et vice-versa."""
        if state == Qt.Checked:
            for label, (cb, key) in self.gbif_checkboxes.items():
                if label != "Tous groupes":
                    cb.setChecked(False)

    def _on_gbif_toggle(self, state):
        """Affiche ou masque les options GBIF."""
        self.gbif_widget.setVisible(state == Qt.Checked)

    def _load_gbif(self, project, root, grp,extent_effectif=None, lyr_clip_effectif=None):
        """Charge les observations GBIF — multi-taxons, pagination complète."""
        import requests
        from qgis.core import (QgsVectorLayer, QgsFeature, QgsGeometry,
                                QgsPointXY, QgsFields, QgsField,
                                QgsLayerTreeLayer, QgsMarkerSymbol,
                                QgsSimpleMarkerSymbolLayer,
                                QgsSingleSymbolRenderer)
        from qgis.PyQt.QtCore import QVariant

        # ── Récupérer les taxons sélectionnés ────────────────────────────
        taxon_keys = []
        tous_coches = self.gbif_checkboxes["Tous groupes"][0].isChecked()

        if not tous_coches:
            for label, (cb, key) in self.gbif_checkboxes.items():
                if cb.isChecked() and key is not None:
                    taxon_keys.append((label, key))
            if not taxon_keys:
                self.lbl_status.setText("⚠ GBIF : sélectionnez au moins un groupe.")
                return
        # Si tous groupes : taxon_keys reste vide → pas de filtre taxon

        # ── Limite ───────────────────────────────────────────────────────
        limit = int(self.combo_gbif_limit.currentText().replace(" ", ""))

        # ── Période ──────────────────────────────────────────────────────
        annee_min = self.txt_gbif_annee_min.text().strip()
        annee_max = self.txt_gbif_annee_max.text().strip()

        # ── Paramètres de base ───────────────────────────────────────────
        base_params = {
            "country"            : "FR",
            "hasCoordinate"      : "true",
            "hasGeospatialIssue" : "false",
        }
        if annee_min and annee_max:
            base_params["year"] = f"{annee_min},{annee_max}"
        elif annee_min:
            base_params["year"] = f"{annee_min},{annee_min}"

        # Filtre mois (index 0 = "—" = pas de filtre)
        mois_min_idx = self.combo_gbif_mois_min.currentIndex()
        mois_max_idx = self.combo_gbif_mois_max.currentIndex()
        if mois_min_idx > 0 and mois_max_idx > 0:
            base_params["month"] = f"{mois_min_idx},{mois_max_idx}"
        elif mois_min_idx > 0:
            base_params["month"] = f"{mois_min_idx},{mois_min_idx}"
        elif mois_max_idx > 0:
            base_params["month"] = f"1,{mois_max_idx}"

        if extent_effectif:
            ext = extent_effectif
            base_params["decimalLongitude"] = f"{ext.xMinimum():.4f},{ext.xMaximum():.4f}"
            base_params["decimalLatitude"]  = f"{ext.yMinimum():.4f},{ext.yMaximum():.4f}"

        # ── Construire la liste des requêtes à faire ─────────────────────
        # Une requête par taxon sélectionné, ou une seule si tous groupes
        requetes = []
        if taxon_keys:
            for label, key in taxon_keys:
                p = dict(base_params)
                p["taxonKey"] = key
                requetes.append((label, p))
        else:
            requetes.append(("Tous groupes", base_params))

        # ── Pagination et collecte ───────────────────────────────────────
        all_results = []
        limit_par_taxon = limit // len(requetes) if len(requetes) > 1 else limit

        for label_req, params in requetes:
            collected = 0
            offset    = 0
            page_size = 300

            self.lbl_status.setText(f"GBIF — chargement {label_req}...")
            QApplication.processEvents()

            while collected < limit_par_taxon:
                params["limit"]  = min(page_size, limit_par_taxon - collected)
                params["offset"] = offset
                try:
                    r = requests.get(GBIF_API, params=params, timeout=20)
                    if r.status_code != 200:
                        break
                    data    = r.json()
                    results = data.get("results", [])
                    if not results:
                        break
                    # Ajouter le label du groupe à chaque observation
                    for obs in results:
                        obs["_groupe_label"] = label_req
                    all_results.extend(results)
                    collected += len(results)
                    if data.get("endOfRecords", True):
                        break
                    offset += page_size
                except Exception:
                    break

        if not all_results:
            self.lbl_status.setText("⚠ GBIF : aucune observation trouvée.")
            return

        # ── Créer la couche mémoire ──────────────────────────────────────
        fields = QgsFields()
        fields.append(QgsField("gbifID",           QVariant.String))
        fields.append(QgsField("espece",           QVariant.String))
        fields.append(QgsField("nom_vernaculaire", QVariant.String))
        fields.append(QgsField("famille",          QVariant.String))
        fields.append(QgsField("classe",           QVariant.String))
        fields.append(QgsField("groupe",           QVariant.String))
        fields.append(QgsField("date_obs",         QVariant.String))
        fields.append(QgsField("annee",            QVariant.Int))
        fields.append(QgsField("source",           QVariant.String))
        fields.append(QgsField("latitude",         QVariant.Double))
        fields.append(QgsField("longitude",        QVariant.Double))

        groupes_label = (
            "Tous groupes" if tous_coches
            else " + ".join(l for l, _ in taxon_keys)
        )
        # Construire la partie période du nom
        if annee_min and annee_max:
            periode_label = f"{annee_min}–{annee_max}"
        elif annee_min:
            periode_label = f"depuis {annee_min}"
        elif annee_max:
            periode_label = f"jusqu'à {annee_max}"
        else:
            periode_label = "toutes périodes"

        # Ajouter les mois au label si filtrés
        MOIS_COURTS = ["", "Jan", "Fév", "Mar", "Avr", "Mai", "Jun",
                       "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"]
        mois_min_idx = self.combo_gbif_mois_min.currentIndex()
        mois_max_idx = self.combo_gbif_mois_max.currentIndex()
        if mois_min_idx > 0 or mois_max_idx > 0:
            m_min = MOIS_COURTS[mois_min_idx] if mois_min_idx > 0 else "Jan"
            m_max = MOIS_COURTS[mois_max_idx] if mois_max_idx > 0 else "Déc"
            periode_label += f" ({m_min}–{m_max})"

        lyr_name = f"GBIF — {groupes_label} — {periode_label}"

        mem_lyr = QgsVectorLayer(
            "Point?crs=EPSG:4326&index=yes", lyr_name, "memory")
        pr = mem_lyr.dataProvider()
        pr.addAttributes(fields)
        mem_lyr.updateFields()

        feats = []
        for obs in all_results:
            lat = obs.get("decimalLatitude")
            lon = obs.get("decimalLongitude")
            if lat is None or lon is None:
                continue
            f = QgsFeature(mem_lyr.fields())
            f.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(lon, lat)))
            f.setAttribute("gbifID",           str(obs.get("gbifID", "")))
            f.setAttribute("espece",           obs.get("species",
                                               obs.get("scientificName", "")))
            f.setAttribute("nom_vernaculaire", obs.get("vernacularName", ""))
            f.setAttribute("famille",          obs.get("family", ""))
            f.setAttribute("classe",           obs.get("class", ""))
            f.setAttribute("groupe",           obs.get("_groupe_label", ""))
            f.setAttribute("date_obs",         obs.get("eventDate", ""))
            f.setAttribute("annee",            obs.get("year"))
            f.setAttribute("source",           obs.get("datasetName", ""))
            f.setAttribute("latitude",         lat)
            f.setAttribute("longitude",        lon)
            feats.append(f)

        pr.addFeatures(feats)
        mem_lyr.updateExtents()

        # ── Symbologie par taxon ─────────────────────────────────────────
        # Couleur et forme par groupe taxonomique
        GBIF_STYLES = {
            "Tous groupes"   : ("#95A5A6", "#607080", QgsSimpleMarkerSymbolLayer.Shape.Circle),
            "🦋 Insectes"    : ("#8E44AD", "#5E2D7D", QgsSimpleMarkerSymbolLayer.Shape.Star),
            "🌿 Plantes"     : ("#27AE60", "#1A7A40", QgsSimpleMarkerSymbolLayer.Shape.Square),
            "🐦 Oiseaux"     : ("#E8A020", "#B06010", QgsSimpleMarkerSymbolLayer.Shape.Circle),
            "🦇 Chiroptères" : ("#2C3E50", "#1A252F", QgsSimpleMarkerSymbolLayer.Shape.Diamond),
            "🐾 Mammifères"  : ("#C0392B", "#8B0000", QgsSimpleMarkerSymbolLayer.Shape.Diamond),
            "🐸 Amphibiens"  : ("#1E8BC3", "#155880", QgsSimpleMarkerSymbolLayer.Shape.Triangle),
            "🦎 Reptiles"    : ("#16A085", "#0E6655", QgsSimpleMarkerSymbolLayer.Shape.Square),
            "🐟 Poissons"    : ("#2980B9", "#1A5276", QgsSimpleMarkerSymbolLayer.Shape.Arrow),
            "🦀 Crustacés"   : ("#E74C3C", "#922B21", QgsSimpleMarkerSymbolLayer.Shape.Pentagon),
            "🐌 Mollusques"  : ("#D35400", "#922B00", QgsSimpleMarkerSymbolLayer.Shape.Pentagon),
            "🍄 Champignons" : ("#E67E22", "#A04000", QgsSimpleMarkerSymbolLayer.Shape.Circle),
            "🌾 Bryophytes"  : ("#2ECC71", "#1A8A4A", QgsSimpleMarkerSymbolLayer.Shape.Cross2),
            "🦟 Arachnides"  : ("#7F8C8D", "#4D5656", QgsSimpleMarkerSymbolLayer.Shape.Star),
        }

        if tous_coches or len(taxon_keys) != 1:
            # Plusieurs taxons ou tous groupes → renderer par règles sur le champ "groupe"
            from qgis.core import QgsRuleBasedRenderer, QgsSymbol
            renderer = QgsRuleBasedRenderer(
                QgsSymbol.defaultSymbol(mem_lyr.geometryType()))
            rule_root = renderer.rootRule()
            rule_root.removeChildAt(0)

            groupes_a_styler = (
                [(l, k) for l, k in GBIF_GROUPES if l != "Tous groupes"]
                if tous_coches
                else taxon_keys
            )

            for label_g, _ in groupes_a_styler:
                fill, stroke, shape = GBIF_STYLES.get(
                    label_g, ("#95A5A6", "#607080",
                              QgsSimpleMarkerSymbolLayer.Shape.Circle))
                sym = QgsMarkerSymbol()
                sym.deleteSymbolLayer(0)
                mk = QgsSimpleMarkerSymbolLayer()
                mk.setShape(shape)
                mk.setSize(1.8);         mk.setSizeUnit(MM)
                mk.setColor(QColor(fill))
                mk.setStrokeColor(QColor(stroke))
                mk.setStrokeWidth(0.15); mk.setStrokeWidthUnit(MM)
                sym.appendSymbolLayer(mk)
                sym.setOpacity(0.75)
                rule = QgsRuleBasedRenderer.Rule(sym)
                rule.setFilterExpression(f'"groupe" = \'{label_g}\'')
                rule.setLabel(label_g)
                rule_root.appendChild(rule)

            # Règle else — observations sans groupe identifié
            sym_else = QgsMarkerSymbol()
            sym_else.deleteSymbolLayer(0)
            mk_else = QgsSimpleMarkerSymbolLayer()
            mk_else.setShape(QgsSimpleMarkerSymbolLayer.Shape.Circle)
            mk_else.setSize(1.4); mk_else.setSizeUnit(MM)
            mk_else.setColor(QColor("#BDC3C7"))
            mk_else.setStrokeColor(QColor("#7F8C8D"))
            mk_else.setStrokeWidth(0.12); mk_else.setStrokeWidthUnit(MM)
            sym_else.appendSymbolLayer(mk_else)
            sym_else.setOpacity(0.55)
            rule_else = QgsRuleBasedRenderer.Rule(sym_else)
            rule_else.setIsElse(True)
            rule_else.setLabel("Autre / inconnu")
            rule_root.appendChild(rule_else)

            mem_lyr.setRenderer(renderer)

        else:
            # Un seul taxon → symbole unique avec sa couleur dédiée
            label_seul = taxon_keys[0][0]
            fill, stroke, shape = GBIF_STYLES.get(
                label_seul, ("#E67E22", "#A04000",
                             QgsSimpleMarkerSymbolLayer.Shape.Circle))
            sym = QgsMarkerSymbol()
            sym.deleteSymbolLayer(0)
            mk = QgsSimpleMarkerSymbolLayer()
            mk.setShape(shape)
            mk.setSize(1.8);         mk.setSizeUnit(MM)
            mk.setColor(QColor(fill))
            mk.setStrokeColor(QColor(stroke))
            mk.setStrokeWidth(0.15); mk.setStrokeWidthUnit(MM)
            sym.appendSymbolLayer(mk)
            sym.setOpacity(0.75)
            mem_lyr.setRenderer(QgsSingleSymbolRenderer(sym))

        project.addMapLayer(mem_lyr, False)
        if grp:
            grp.addChildNode(QgsLayerTreeLayer(mem_lyr))
        else:
            root.addLayer(mem_lyr)

        self.lbl_status.setText(
            f"✅ GBIF : {len(feats):,} observations — {groupes_label}")

    def _on_export_toggle(self, state):
        """Affiche ou masque les options d'export Excel."""
        self.export_widget.setVisible(state == Qt.Checked)

    def _browse_export_path(self):
        """Sélectionner le dossier de sortie pour l'export Excel."""
        from qgis.PyQt.QtWidgets import QFileDialog
        dossier = QFileDialog.getExistingDirectory(
            self, "Choisir le dossier de sortie", "")
        if dossier:
            self.txt_export_path.setText(dossier)

    def _export_excel(self, layers_loaded, territoire_label, buffer_label):
        """Exporte chaque couche chargée dans un onglet Excel."""
        import re, unicodedata, os
        from datetime import datetime

        layers_loaded = [
            (lyr_name, lyr)
            for lyr_name, lyr in layers_loaded
            if not exclude_from_excel(lyr_name)
        ]

        if not layers_loaded:
            self.lbl_status.setText(
                "⚠ Export Excel : aucune couche exportable hors cadastre."
            )
            return

        try:
            import openpyxl
        except ImportError:
            try:
                import subprocess, sys
                subprocess.run([sys.executable, "-m", "pip", "install",
                                "openpyxl", "--break-system-packages"],
                               capture_output=True)
                import openpyxl
            except Exception:
                self.lbl_status.setText("❌ Export Excel : impossible d'installer openpyxl.")
                return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # ── Champs techniques à exclure systématiquement ──────────────────
        CHAMPS_EXCLUS = {
            "gml_id", "gml_boundedby", "fid", "objectid", "ogc_fid",
            "id", "shape_area", "shape_length", "shape_leng",
            "the_geom", "geom", "geometry", "wkb_geometry",
        }

        def clean_sheet_name(raw):
            s = unicodedata.normalize('NFKD', raw)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            s = re.sub(r'[^\x20-\x7E]', '', s)
            for c in r'\/:*?[]':
                s = s.replace(c, "_")
            s = re.sub(r'\s+', ' ', s).strip(" -—_")
            return s[:31] or "Feuille"

        def clean_val(val):
            if val is None or (hasattr(val, 'isNull') and val.isNull()):
                return ""
            if isinstance(val, QDateTime):
                return val.toString("dd/MM/yyyy HH:mm") if val.isValid() else ""
            if isinstance(val, QDate):
                return val.toString("dd/MM/yyyy") if val.isValid() else ""
            if isinstance(val, bool):
                return val
            if isinstance(val, int):
                return val
            if isinstance(val, float):
                return round(val, 4)
            return str(val)

        dossier = self._excel_export_path
        if not dossier:
            self.lbl_status.setText("❌ Export Excel : choisissez un dossier de sortie.")
            return

        nom_base = f"{territoire_label}_{buffer_label}"
        nom_base = unicodedata.normalize('NFKD', nom_base)
        nom_base = ''.join(c for c in nom_base if unicodedata.category(c) != 'Mn')
        nom_base = re.sub(r'[^\x20-\x7E]', '', nom_base)
        for c in r'\/:*?"<>|':
            nom_base = nom_base.replace(c, "_")
        nom_base = re.sub(r'\s+', '_', nom_base).strip("_")
        chemin_excel = os.path.join(dossier, f"{nom_base}.xlsx")

        # ── Styles ────────────────────────────────────────────────────────
        HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
        HDR_FILL   = PatternFill("solid", fgColor="10B981")   # vert accent
        HDR_ALIGN  = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
        ROW_ODD    = PatternFill("solid", fgColor="F0FDF8")   # vert très pâle
        ROW_EVEN   = PatternFill("solid", fgColor="FFFFFF")
        CELL_ALIGN = Alignment(vertical="top", wrap_text=False)
        RECAP_HDR_FILL = PatternFill("solid", fgColor="1E2235")
        RECAP_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

        wb = Workbook()
        wb.remove(wb.active)

        recap_rows = []  # (nom_couche, nb_entités, onglet)

        for lyr_name, lyr in layers_loaded:
            self.lbl_status.setText(f"Excel — mise en forme : {lyr_name}...")
            QApplication.processEvents()

            # ── Filtrer les champs parasites ──────────────────────────────
            all_fields = [f.name() for f in lyr.fields()]
            fields_kept = [f for f in all_fields
                           if f.lower() not in CHAMPS_EXCLUS]

            # ── Récupérer toutes les entités ──────────────────────────────
            feats = list(lyr.getFeatures())
            nb = len(feats)

            # ── Supprimer les colonnes entièrement vides ──────────────────
            non_null = set()
            for feat in feats:
                for f in fields_kept:
                    v = feat[f]
                    if v is not None and not (hasattr(v, 'isNull') and v.isNull()) \
                            and str(v).strip() != "":
                        non_null.add(f)
            fields_kept = [f for f in fields_kept if f in non_null]

            sheet_name = clean_sheet_name(lyr_name)
            ws = wb.create_sheet(title=sheet_name)

            # ── En-têtes ──────────────────────────────────────────────────
            col_widths = {}
            for col, field in enumerate(fields_kept, 1):
                cell = ws.cell(row=1, column=col, value=field)
                cell.font      = HDR_FONT
                cell.fill      = HDR_FILL
                cell.alignment = HDR_ALIGN
                col_widths[col] = max(10, len(field) + 2)

            # ── Données ───────────────────────────────────────────────────
            for row_idx, feat in enumerate(feats, 2):
                row_fill = ROW_ODD if row_idx % 2 == 0 else ROW_EVEN
                for col_idx, field in enumerate(fields_kept, 1):
                    val = clean_val(feat[field])
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill      = row_fill
                    cell.alignment = CELL_ALIGN
                    # Ajuster la largeur de colonne sur les 200 premières lignes
                    if row_idx <= 202:
                        w = min(50, max(col_widths[col_idx],
                                        len(str(val)) + 2 if val != "" else 0))
                        col_widths[col_idx] = w

            # ── Appliquer les largeurs calculées ──────────────────────────
            for col, w in col_widths.items():
                ws.column_dimensions[get_column_letter(col)].width = w

            # ── Figer 1re ligne + filtres automatiques ────────────────────
            ws.freeze_panes = "A2"
            if fields_kept:
                ws.auto_filter.ref = (
                    f"A1:{get_column_letter(len(fields_kept))}1"
                )

            # Hauteur de ligne standard pour les données
            for row in range(2, nb + 2):
                ws.row_dimensions[row].height = 15

            recap_rows.append((lyr_name, nb, sheet_name))

        # ── Onglet récapitulatif (en première position) ────────────────────
        ws_recap = wb.create_sheet(title="Récapitulatif", index=0)
        recap_headers = ["Couche", "Entités", "Onglet", "Territoire", "Date export"]
        for col, h in enumerate(recap_headers, 1):
            cell = ws_recap.cell(row=1, column=col, value=h)
            cell.font      = RECAP_HDR_FONT
            cell.fill      = RECAP_HDR_FILL
            cell.alignment = HDR_ALIGN

        date_export = datetime.now().strftime("%d/%m/%Y %H:%M")
        for row_idx, (nom, nb, onglet) in enumerate(recap_rows, 2):
            row_fill = ROW_ODD if row_idx % 2 == 0 else ROW_EVEN
            for col, val in enumerate(
                [nom, nb, onglet, territoire_label, date_export], 1
            ):
                cell = ws_recap.cell(row=row_idx, column=col, value=val)
                cell.fill      = row_fill
                cell.alignment = CELL_ALIGN

        recap_widths = [40, 10, 28, 28, 18]
        for col, w in enumerate(recap_widths, 1):
            ws_recap.column_dimensions[get_column_letter(col)].width = w
        ws_recap.freeze_panes = "A2"
        ws_recap.auto_filter.ref = f"A1:{get_column_letter(len(recap_headers))}1"

        if not wb.sheetnames:
            self.lbl_status.setText("⚠ Export Excel : aucune couche à exporter.")
            return

        try:
            wb.save(chemin_excel)
            self.lbl_status.setText(f"✅ Excel exporté : {nom_base}.xlsx")
        except PermissionError:
            self._msgbox("critical", "Fichier inaccessible",
                f"Impossible d'enregistrer le fichier :\n\n{chemin_excel}\n\n"
                f"Le fichier est peut-être déjà ouvert dans Excel.\n"
                f"Fermez-le puis relancez l'export.")
            self.lbl_status.setText("❌ Export Excel annulé — fichier déjà ouvert.")
        except Exception as e:
            self._msgbox("critical", "Erreur d'export",
                f"Une erreur inattendue s'est produite lors de l'export :\n\n{str(e)}")
            self.lbl_status.setText(f"❌ Export Excel échoué : {str(e)[:60]}")

    def _export_excel_from_project(self):
        """Export Excel — demande le dossier si nécessaire, recharge et exporte."""
        from qgis.core import QgsVectorLayer
        from qgis.PyQt.QtWidgets import QFileDialog

        if not self._excel_export_path:
            dossier = QFileDialog.getExistingDirectory(
                self, "Choisir le dossier de sortie pour l'export Excel", "")
            if not dossier:
                return
            self._excel_export_path = dossier

        selected = [nom for nom, cb in self.checkboxes.items() if cb.isChecked()]
        if not selected:
            self._msgbox("warning", "Attention",
                "Cochez au moins une couche dans la section 2 avant d'exporter.")
            return

        if self.combo_scale.currentText() != "France entière" and \
                self._territory_extent is None:
            self._msgbox("warning", "Attention",
                "Recherchez d'abord un territoire avant d'exporter.")
            return

        self.btn_export_excel.setEnabled(False)
        self.lbl_status.setText("Préparation de l'export Excel...")
        QApplication.processEvents()

        # ── Buffer si coché ───────────────────────────────────────────────
        lyr_clip_effectif = self._territory_lyr
        extent_effectif   = self._territory_extent

        if self.chk_buffer.isChecked() and self._territory_lyr:
            self.lbl_status.setText("Calcul du buffer pour l'export...")
            QApplication.processEvents()
            lyr_buf = self._apply_buffer(self._territory_lyr)
            if lyr_buf:
                lyr_clip_effectif = lyr_buf
                extent_effectif   = lyr_buf.extent()

        catalogue_dict = {nom: (typename, serveur, fill, stroke)
                          for _, nom, typename, serveur, fill, stroke in CATALOGUE}

        layers_for_export = []

        for nom in selected:
            if exclude_from_excel(nom):
                self.lbl_status.setText(f"Export — {nom} : exclu de l'export Excel")
                QApplication.processEvents()
                continue

            self.lbl_status.setText(f"Export — chargement : {nom}...")
            QApplication.processEvents()

            typename, serveur, fill, stroke = catalogue_dict[nom]

            # Construire l'URI WFS avec bbox
            if extent_effectif:
                ext = extent_effectif
                mx  = ext.width()  * 0.005
                my  = ext.height() * 0.005
                bbox_str = (f"{ext.xMinimum()-mx},{ext.yMinimum()-my},"
                            f"{ext.xMaximum()+mx},{ext.yMaximum()+my}")
                uri, filtre_expr, is_gpu = build_wfs_uri(serveur, typename, bbox_str)
            else:
                uri, filtre_expr, is_gpu = build_wfs_uri(serveur, typename, None)


            # Parcelles cadastrales : API Carto JSON (rapide, endpoint confirmé)
            if typename == "CADASTRALPARCELS.PARCELLAIRE_EXPRESS:parcelle" and lyr_clip_effectif:
                code_cadastre = (
                    self._territory_code
                    if self.combo_scale.currentText() == "Commune"
                    else None
                )
                lyr = load_cadastre_api(
                    API_CARTO_CADASTRE, lyr_clip_effectif, nom,
                    self.lbl_status, QApplication, code_insee=code_cadastre
                )
                if not lyr or not layer_has_features(lyr):
                    self.lbl_status.setText(f"Export — {nom} : aucune parcelle trouvée")
                    QApplication.processEvents()
                    continue
                layers_for_export.append((nom, lyr))
                continue

            # Couches GPU : chargement spécial par code INSEE
            if is_gpu and lyr_clip_effectif:
                lyr = load_gpu_layer(typename, filtre_expr, lyr_clip_effectif,
                                     nom, self.lbl_status, QApplication)

                if not lyr or not layer_has_features(lyr):
                    self.lbl_status.setText(f"Export — {nom} : aucune donnée, ignoré")
                    QApplication.processEvents()
                    continue

                layers_for_export.append((nom, lyr))
                continue

            lyr = QgsVectorLayer(uri, nom, "WFS")

            if not lyr.isValid():
                self.lbl_status.setText(f"Export — {nom} : couche invalide")
                QApplication.processEvents()
                continue

            if is_sup_layer(typename):
                self.lbl_status.setText(f"SUP — préparation : {nom}...")
                QApplication.processEvents()
                lyr = materialize_layer(lyr, nom)
                try:
                    self.lbl_status.setText(f"SUP — correction géométries : {nom}...")
                    QApplication.processEvents()
                    lyr = processing.run("native:fixgeometries", {
                        'INPUT':  lyr,
                        'OUTPUT': 'memory:'
                    })['OUTPUT']
                except Exception:
                    pass

            # ── Étape 1 : clip spatial rapide ─────────────────────────────
            # On évite fixgeometries par défaut : très coûteux sur les WFS.
            # On ne l'utilise qu'en secours si le clip direct échoue.
            if lyr_clip_effectif:
                try:
                    if self.chk_intersect_only.isChecked():
                        res = processing.run("native:extractbylocation", {
                            'INPUT':     lyr,
                            'PREDICATE': [0],
                            'INTERSECT': lyr_clip_effectif,
                            'OUTPUT':    'memory:'
                        })
                    else:
                        res = processing.run("native:clip", {
                            'INPUT':   lyr,
                            'OVERLAY': lyr_clip_effectif,
                            'OUTPUT':  'memory:'
                        })

                    lyr_filtre = res['OUTPUT']
                    lyr_filtre.setName(nom)
                    lyr = lyr_filtre

                except Exception:
                    try:
                        lyr_fixed = processing.run("native:fixgeometries", {
                            'INPUT':  lyr,
                            'OUTPUT': 'memory:'
                        })['OUTPUT']

                        if self.chk_intersect_only.isChecked():
                            res = processing.run("native:extractbylocation", {
                                'INPUT':     lyr_fixed,
                                'PREDICATE': [0],
                                'INTERSECT': lyr_clip_effectif,
                                'OUTPUT':    'memory:'
                            })
                        else:
                            res = processing.run("native:clip", {
                                'INPUT':   lyr_fixed,
                                'OVERLAY': lyr_clip_effectif,
                                'OUTPUT':  'memory:'
                            })

                        lyr_filtre = res['OUTPUT']
                        lyr_filtre.setName(nom)
                        lyr = lyr_filtre

                    except Exception as e:
                        self.lbl_status.setText(f"⚠ Clip échoué : {str(e)[:60]}")
                        QApplication.processEvents()

            # ── Étape 2 : filtre attributaire APRÈS le clip ───────────────
            if filtre_expr and layer_has_features(lyr):
                from qgis.core import QgsFeatureRequest, QgsExpression
                expr    = QgsExpression(filtre_expr)
                request = QgsFeatureRequest(expr)
                feats   = list(lyr.getFeatures(request))
                crs_id  = lyr.crs().authid()
                fields  = lyr.fields().toList()
                lyr_mem = QgsVectorLayer(
                    f"Polygon?crs={crs_id}", nom, "memory")
                pr_mem  = lyr_mem.dataProvider()
                pr_mem.addAttributes(fields)
                lyr_mem.updateFields()
                pr_mem.addFeatures(feats)
                lyr_mem.updateExtents()
                lyr = lyr_mem

            if layer_has_features(lyr):
                layers_for_export.append((nom, lyr))

        if not layers_for_export:
            self._msgbox("warning", "Export Excel",
                "Aucune donnée trouvée sur ce territoire pour les couches sélectionnées.")
            self.btn_export_excel.setEnabled(True)
            return

        # Labels pour le nom du fichier
        territoire_label = self.txt_search.text().strip() or "France"
        scale = self.combo_scale.currentText()
        territoire_label = f"{scale}_{territoire_label}"

        if self.chk_buffer.isChecked():
            taille = self.txt_buffer_size.text().strip()
            unite  = self.combo_buffer_unit.currentText()
            buffer_label = f"buffer_{taille}_{unite}"
        else:
            buffer_label = "sans_buffer"

        self._export_excel(layers_for_export, territoire_label, buffer_label)
        self.btn_export_excel.setEnabled(True)

    def _on_gpkg_toggle(self, state):
        """Affiche ou masque le sélecteur de GeoPackage."""
        self.gpkg_widget.setVisible(state == Qt.Checked)

    def _browse_gpkg_path(self):
        """Sélectionner ou créer un fichier GeoPackage de sortie."""
        from qgis.PyQt.QtWidgets import QFileDialog
        chemin, _ = QFileDialog.getSaveFileName(
            self, "Choisir ou créer un GeoPackage", "",
            "GeoPackage (*.gpkg)")
        if chemin:
            if not chemin.endswith(".gpkg"):
                chemin += ".gpkg"
            self.txt_gpkg_path.setText(chemin)

    def _save_layer_to_gpkg(self, lyr, gpkg_path, first):
        """Sauvegarde une couche dans le GeoPackage et recharge depuis le GPKG
        pour conserver la symbologie et remplacer la couche temporaire."""
        from qgis.core import (QgsVectorFileWriter, QgsCoordinateTransformContext,
                                QgsProject, QgsLayerTreeLayer)
        import os

        options = QgsVectorFileWriter.SaveVectorOptions()
        options.driverName   = "GPKG"
        options.layerName    = lyr.name()[:60]  # nom de couche GPKG
        options.fileEncoding = "UTF-8"
        options.actionOnExistingFile = (
            QgsVectorFileWriter.CreateOrOverwriteFile if first
            else QgsVectorFileWriter.CreateOrOverwriteLayer
        )

        err, msg, _, _ = QgsVectorFileWriter.writeAsVectorFormatV3(
            lyr, gpkg_path, QgsCoordinateTransformContext(), options)

        if err != 0:
            return None, f"Erreur GPKG pour {lyr.name()} : {msg}"

        # Recharger depuis le GPKG pour avoir une couche persistante
        layer_name = options.layerName
        uri = f"{gpkg_path}|layername={layer_name}"
        lyr_gpkg = QgsVectorLayer(uri, lyr.name(), "ogr")

        if not lyr_gpkg.isValid():
            return None, f"Couche GPKG invalide : {lyr.name()}"

        # Copier la symbologie de la couche temporaire vers la couche GPKG
        lyr_gpkg.setRenderer(lyr.renderer().clone())
        lyr_gpkg.triggerRepaint()

        return lyr_gpkg, None

    def _export_to_gpkg(self):
        """Recharge les couches cochées, filtre par zone et exporte dans un GeoPackage."""
        from qgis.PyQt.QtWidgets import QFileDialog
        from qgis.core import (QgsVectorLayer, QgsVectorFileWriter,
                                QgsCoordinateTransformContext)
        import re, unicodedata

        selected = [nom for nom, cb in self.checkboxes.items() if cb.isChecked()]
        if not selected:
            self._msgbox("warning", "Attention",
                "Cochez au moins une couche dans la section 2.")
            return

        if self.combo_scale.currentText() != "France entière" and \
                self._territory_extent is None:
            self._msgbox("warning", "Attention",
                "Recherchez d'abord un territoire.")
            return

        # Choisir le fichier de sortie
        chemin_gpkg, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le GeoPackage", "",
            "GeoPackage (*.gpkg)")
        if not chemin_gpkg:
            return
        if not chemin_gpkg.endswith(".gpkg"):
            chemin_gpkg += ".gpkg"

        self.btn_save_gpkg.setEnabled(False)
        self.lbl_status.setText("Préparation de l'export GeoPackage...")
        QApplication.processEvents()

        # Buffer si coché
        lyr_clip_effectif = self._territory_lyr
        extent_effectif   = self._territory_extent

        if self.chk_buffer.isChecked() and self._territory_lyr:
            self.lbl_status.setText("Calcul du buffer...")
            QApplication.processEvents()
            lyr_buf = self._apply_buffer(self._territory_lyr)
            if lyr_buf:
                lyr_clip_effectif = lyr_buf
                extent_effectif   = lyr_buf.extent()

        catalogue_dict = {nom: (typename, serveur, fill, stroke)
                          for _, nom, typename, serveur, fill, stroke in CATALOGUE}

        project = QgsProject.instance()
        root    = project.layerTreeRoot()
        if self.chk_group.isChecked():
            territoire = self.txt_search.text().strip() or "France"
            scale = self.combo_scale.currentText()
            grp = QgsLayerTreeGroup(f"BioBlio — {scale} — {territoire}")
            grp.setExpanded(True)
            root.insertChildNode(0, grp)
        else:
            grp = None

        first = True
        nb_ok = 0

        for nom in selected:
            self.lbl_status.setText(f"GPKG — chargement : {nom}...")
            QApplication.processEvents()

            typename, serveur, fill, stroke = catalogue_dict[nom]

            # URI WFS
            if extent_effectif:
                ext = extent_effectif
                mx  = ext.width()  * 0.005
                my  = ext.height() * 0.005
                bbox_str = (f"{ext.xMinimum()-mx},{ext.yMinimum()-my},"
                            f"{ext.xMaximum()+mx},{ext.yMaximum()+my}")
                uri, filtre_expr, is_gpu = build_wfs_uri(serveur, typename, bbox_str)
            else:
                uri, filtre_expr, is_gpu = build_wfs_uri(serveur, typename, None)

            # Couches GPU : chargement spécial par code INSEE
            if is_gpu and lyr_clip_effectif:
                lyr = load_gpu_layer(typename, filtre_expr, lyr_clip_effectif,
                                     nom, self.lbl_status, QApplication)

                if not lyr or not layer_has_features(lyr):
                    self.lbl_status.setText(f"{nom} : aucune donnée GPU trouvée")
                    QApplication.processEvents()
                    continue

                # Clip post-chargement sur la zone effective
                self.lbl_status.setText(f"GPU — filtrage spatial : {nom}...")
                QApplication.processEvents()
                try:
                    if self.chk_intersect_only.isChecked():
                        res = processing.run("native:extractbylocation", {
                            'INPUT':     lyr,
                            'PREDICATE': [0],
                            'INTERSECT': lyr_clip_effectif,
                            'OUTPUT':    'memory:'
                        })
                    else:
                        res = processing.run("native:clip", {
                            'INPUT':   lyr,
                            'OVERLAY': lyr_clip_effectif,
                            'OUTPUT':  'memory:'
                        })
                    lyr_clipped = res['OUTPUT']
                    lyr_clipped.setName(nom)
                    if layer_has_features(lyr_clipped):
                        lyr = lyr_clipped
                except Exception:
                    pass

                if not layer_has_features(lyr):
                    QApplication.processEvents()
                    continue

            else:
                lyr = QgsVectorLayer(uri, nom, "WFS")
                if not lyr.isValid():
                    continue

                if is_sup_layer(typename):
                    self.lbl_status.setText(f"SUP — préparation : {nom}...")
                    QApplication.processEvents()
                    lyr = materialize_layer(lyr, nom)
                    try:
                        self.lbl_status.setText(f"SUP — correction géométries : {nom}...")
                        QApplication.processEvents()
                        lyr = processing.run("native:fixgeometries", {
                            'INPUT':  lyr,
                            'OUTPUT': 'memory:'
                        })['OUTPUT']
                    except Exception:
                        pass

            # ── Étape 1 : clip spatial rapide ─────────────────────────────
            # On évite fixgeometries par défaut : très coûteux sur les WFS.
            # On ne l'utilise qu'en secours si le clip direct échoue.
            if lyr_clip_effectif:
                try:
                    if self.chk_intersect_only.isChecked():
                        res = processing.run("native:extractbylocation", {
                            'INPUT':     lyr,
                            'PREDICATE': [0],
                            'INTERSECT': lyr_clip_effectif,
                            'OUTPUT':    'memory:'
                        })
                    else:
                        res = processing.run("native:clip", {
                            'INPUT':   lyr,
                            'OVERLAY': lyr_clip_effectif,
                            'OUTPUT':  'memory:'
                        })

                    lyr_f = res['OUTPUT']
                    lyr_f.setName(nom)
                    lyr = lyr_f

                except Exception:
                    try:
                        lyr_fixed = processing.run("native:fixgeometries", {
                            'INPUT':  lyr,
                            'OUTPUT': 'memory:'
                        })['OUTPUT']

                        if self.chk_intersect_only.isChecked():
                            res = processing.run("native:extractbylocation", {
                                'INPUT':     lyr_fixed,
                                'PREDICATE': [0],
                                'INTERSECT': lyr_clip_effectif,
                                'OUTPUT':    'memory:'
                            })
                        else:
                            res = processing.run("native:clip", {
                                'INPUT':   lyr_fixed,
                                'OVERLAY': lyr_clip_effectif,
                                'OUTPUT':  'memory:'
                            })

                        lyr_f = res['OUTPUT']
                        lyr_f.setName(nom)
                        lyr = lyr_f

                    except Exception as e:
                        self.lbl_status.setText(f"⚠ Clip échoué : {str(e)[:60]}")
                        QApplication.processEvents()

            # ── Étape 2 : filtre attributaire APRÈS le clip ───────────────
            if filtre_expr and layer_has_features(lyr):
                from qgis.core import QgsFeatureRequest, QgsExpression
                expr    = QgsExpression(filtre_expr)
                request = QgsFeatureRequest(expr)
                feats   = list(lyr.getFeatures(request))
                crs_id  = lyr.crs().authid()
                fields  = lyr.fields().toList()
                lyr_mem = QgsVectorLayer(
                    f"Polygon?crs={crs_id}", nom, "memory")
                pr_mem  = lyr_mem.dataProvider()
                pr_mem.addAttributes(fields)
                lyr_mem.updateFields()
                pr_mem.addFeatures(feats)
                lyr_mem.updateExtents()
                lyr = lyr_mem

            if not layer_has_features(lyr):
                continue

            # Appliquer la symbologie
            if self.chk_style.isChecked():
                apply_style(lyr, fill, stroke)

            # Nom de layer propre pour le GPKG
            layer_name = unicodedata.normalize('NFKD', nom)
            layer_name = ''.join(c for c in layer_name
                                 if unicodedata.category(c) != 'Mn')
            import re
            layer_name = re.sub(r'[^\w\s-]', '', layer_name).strip()
            layer_name = re.sub(r'\s+', '_', layer_name)[:60]

            self.lbl_status.setText(f"GPKG — écriture : {nom}...")
            QApplication.processEvents()

            options = QgsVectorFileWriter.SaveVectorOptions()
            options.driverName   = "GPKG"
            options.layerName    = layer_name
            options.fileEncoding = "UTF-8"
            options.actionOnExistingFile = (
                QgsVectorFileWriter.CreateOrOverwriteFile if first
                else QgsVectorFileWriter.CreateOrOverwriteLayer
            )

            err, msg, _, _ = QgsVectorFileWriter.writeAsVectorFormatV3(
                lyr, chemin_gpkg, QgsCoordinateTransformContext(), options)

            if err == 0:
                # Charger depuis le GPKG avec symbologie et ajouter au projet
                uri_gpkg = f"{chemin_gpkg}|layername={layer_name}"
                lyr_gpkg = QgsVectorLayer(uri_gpkg, nom, "ogr")
                if lyr_gpkg.isValid():
                    lyr_gpkg.setRenderer(lyr.renderer().clone())
                    project.addMapLayer(lyr_gpkg, False)
                    if grp:
                        grp.addChildNode(QgsLayerTreeLayer(lyr_gpkg))
                    else:
                        root.addLayer(lyr_gpkg)
                first = False
                nb_ok += 1
            else:
                self.lbl_status.setText(f"⚠ GPKG erreur : {nom} — {msg}")

        if nb_ok > 0:
            self.lbl_status.setText(
                f"✅ GeoPackage exporté : {nb_ok} couche(s) → {chemin_gpkg}")
            self._msgbox("info", "Export GeoPackage",
                f"{nb_ok} couche(s) exportée(s) dans :\n{chemin_gpkg}\n\n"
                f"Les couches ont été rechargées depuis le GeoPackage dans le projet.")
        else:
            self.lbl_status.setText("⚠ Aucune couche exportée.")

        self.btn_save_gpkg.setEnabled(True)

    def _check_all(self, state):
        for cb in self.checkboxes.values():
            if cb.isEnabled():
                cb.setChecked(state)

    def _apply_buffer(self, lyr):
        """Applique un buffer sur la couche territoire et retourne la couche bufférisée."""
        import processing

        try:
            taille = float(self.txt_buffer_size.text().strip().replace(",", "."))
        except ValueError:
            self.lbl_result.setText("❌ Taille de buffer invalide.")
            self.lbl_result.setStyleSheet("color:red; font-style:italic;")
            return None

        unite = self.combo_buffer_unit.currentText()

        # Reprojeter en Lambert 93 pour un buffer métrique précis
        res_proj = processing.run("native:reprojectlayer", {
            'INPUT':      lyr,
            'TARGET_CRS': 'EPSG:2154',
            'OUTPUT':     'memory:'
        })
        lyr_l93 = res_proj['OUTPUT']

        # Convertir en mètres
        distance_m = taille * 1000 if unite == "Kilomètres" else taille

        # Buffer
        res_buf = processing.run("native:buffer", {
            'INPUT':        lyr_l93,
            'DISTANCE':     distance_m,
            'SEGMENTS':     16,
            'END_CAP_STYLE': 0,
            'JOIN_STYLE':   0,
            'MITER_LIMIT':  2,
            'DISSOLVE':     True,
            'OUTPUT':       'memory:'
        })
        lyr_buf_l93 = res_buf['OUTPUT']

        # Reprojeter en WGS84 pour cohérence avec les WFS
        res_wgs = processing.run("native:reprojectlayer", {
            'INPUT':      lyr_buf_l93,
            'TARGET_CRS': 'EPSG:4326',
            'OUTPUT':     'memory:'
        })
        return res_wgs['OUTPUT']

    def _load_data(self):
        selected = [nom for nom, cb in self.checkboxes.items() if cb.isChecked()]
        if not selected and not self.chk_gbif.isChecked():
            self._msgbox("warning", "Attention",
                "Cochez au moins une couche ou activez le chargement GBIF.")
            return
        if self.combo_scale.currentText() != "France entière" and self._territory_extent is None:
            self._msgbox("warning", "Attention", "Recherchez d'abord un territoire.")
            return

        _heavy = {"🧾 Parcelles cadastrales", "🏠 Bâtiments cadastraux"}
        _heavy_selected = _heavy.intersection(selected)
        if _heavy_selected and self.combo_scale.currentText() != "Commune":
            self._msgbox(
                "warning",
                "Chargement potentiellement long",
                "⚠  Les couches suivantes peuvent nécessiter plusieurs minutes "
                "sur une zone plus grande qu'une commune :\n\n"
                + "\n".join(f"  • {n}" for n in sorted(_heavy_selected))
                + "\n\nLe chargement va démarrer — merci de patienter."
            )

        self.btn_load.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setMaximum(len(selected))
        self.progress.setValue(0)

        project = QgsProject.instance()
        root    = project.layerTreeRoot()

        if self.chk_group.isChecked():
            territoire = self.txt_search.text().strip() or "France"
            scale = self.combo_scale.currentText()
            grp = QgsLayerTreeGroup(f"BioBlio — {scale} — {territoire}")
            grp.setExpanded(True)
            root.addChildNode(grp)
        else:
            grp = None

        catalogue_dict = {nom: (typename, serveur, fill, stroke)
                          for _, nom, typename, serveur, fill, stroke in CATALOGUE}

        urbanisme_names = {
            nom for cat, nom, *_ in CATALOGUE
            if "Urbanisme" in cat
        }
        urbanisme_selected = [nom for nom in selected if nom in urbanisme_names]
        urbanisme_loaded = []

        from qgis.PyQt.QtWidgets import QApplication

        # ── Appliquer le buffer si coché ─────────────────────────────────
        lyr_clip_effectif = self._territory_lyr
        extent_effectif   = self._territory_extent

        if self.chk_buffer.isChecked() and self._territory_lyr:
            self.lbl_status.setText("Calcul du buffer...")
            QApplication.processEvents()
            lyr_buf = self._apply_buffer(self._territory_lyr)
            if lyr_buf:
                lyr_clip_effectif = lyr_buf
                extent_effectif   = lyr_buf.extent()
                taille = self.txt_buffer_size.text().strip()
                unite  = self.combo_buffer_unit.currentText()
                self.lbl_status.setText(f"Buffer {taille} {unite} appliqué ✅")
                QApplication.processEvents()
        
        layers_loaded = []  # liste des couches chargées pour l'export Excel
        layers_empty = []   # couches sélectionnées mais sans donnée finale

        for i, nom in enumerate(selected):
            self.lbl_status.setText(f"Chargement : {nom}...")
            QApplication.processEvents()

            typename, serveur, fill, stroke = catalogue_dict[nom]

            if extent_effectif:
                ext = extent_effectif
                mx  = ext.width()  * 0.005
                my  = ext.height() * 0.005
                bbox_str = (f"{ext.xMinimum()-mx},{ext.yMinimum()-my},"
                            f"{ext.xMaximum()+mx},{ext.yMaximum()+my}")
                uri, filtre_expr, is_gpu = build_wfs_uri(serveur, typename, bbox_str)
            else:
                uri, filtre_expr, is_gpu = build_wfs_uri(serveur, typename, None)

            # Parcelles cadastrales : API Carto JSON (rapide, endpoint confirmé)
            if typename == "CADASTRALPARCELS.PARCELLAIRE_EXPRESS:parcelle" and lyr_clip_effectif:
                code_cadastre = (
                    self._territory_code
                    if self.combo_scale.currentText() == "Commune"
                    else None
                )
                lyr = load_cadastre_api(
                    API_CARTO_CADASTRE, lyr_clip_effectif, nom,
                    self.lbl_status, QApplication, code_insee=code_cadastre
                )
                if not lyr or not layer_has_features(lyr):
                    self.lbl_status.setText(f"{nom} : aucune parcelle trouvée")
                    layers_empty.append(nom)
                    self.progress.setValue(i + 1)
                    QApplication.processEvents()
                    continue
                if self.chk_style.isChecked():
                    apply_style(lyr, fill, stroke)
                project.addMapLayer(lyr, False)
                if grp:
                    grp.addChildNode(QgsLayerTreeLayer(lyr))
                else:
                    root.addLayer(lyr)
                layers_loaded.append((nom, lyr))
                self.progress.setValue(i + 1)
                continue

            # Couches GPU : chargement spécial par code INSEE
            if is_gpu and lyr_clip_effectif:
                lyr = load_gpu_layer(typename, filtre_expr, lyr_clip_effectif,
                                     nom, self.lbl_status, QApplication)

                if not lyr or not layer_has_features(lyr):
                    self.lbl_status.setText(f"{nom} : aucune donnée GPU trouvée")
                    layers_empty.append(nom)
                    self.progress.setValue(i + 1)
                    QApplication.processEvents()
                    continue

                # Clip post-chargement sur la zone effective
                self.lbl_status.setText(f"GPU — filtrage spatial : {nom}...")
                QApplication.processEvents()
                try:
                    if self.chk_intersect_only.isChecked():
                        res = processing.run("native:extractbylocation", {
                            'INPUT':     lyr,
                            'PREDICATE': [0],
                            'INTERSECT': lyr_clip_effectif,
                            'OUTPUT':    'memory:'
                        })
                    else:
                        res = processing.run("native:clip", {
                            'INPUT':   lyr,
                            'OVERLAY': lyr_clip_effectif,
                            'OUTPUT':  'memory:'
                        })
                    lyr_clipped = res['OUTPUT']
                    lyr_clipped.setName(nom)
                    if layer_has_features(lyr_clipped):
                        lyr = lyr_clipped
                except Exception:
                    pass

                if not layer_has_features(lyr):
                    self.lbl_status.setText(f"{nom} : aucune donnée dans la zone")
                    layers_empty.append(nom)
                    self.progress.setValue(i + 1)
                    QApplication.processEvents()
                    continue

                if self.chk_style.isChecked():
                    apply_style(lyr, fill, stroke)

                project.addMapLayer(lyr, False)

                if grp:
                    grp.addChildNode(QgsLayerTreeLayer(lyr))
                else:
                    root.addLayer(lyr)

                layers_loaded.append((nom, lyr))

                if nom in urbanisme_names:
                    urbanisme_loaded.append(nom)

                self.progress.setValue(i + 1)
                continue

            lyr = QgsVectorLayer(uri, nom, "WFS")

            if not lyr.isValid():
                self.lbl_status.setText(f"{nom} : couche invalide")
                layers_empty.append(nom)
                self.progress.setValue(i + 1)
                QApplication.processEvents()
                continue

            if is_sup_layer(typename):
                self.lbl_status.setText(f"SUP — préparation : {nom}...")
                QApplication.processEvents()
                lyr = materialize_layer(lyr, nom)

                if lyr_clip_effectif:
                    # Correction géométries avant clip (SUP a souvent des géométries invalides)
                    lyr_to_clip = lyr
                    try:
                        self.lbl_status.setText(f"SUP — correction géométries : {nom}...")
                        QApplication.processEvents()
                        lyr_to_clip = processing.run("native:fixgeometries", {
                            'INPUT':  lyr,
                            'OUTPUT': 'memory:'
                        })['OUTPUT']
                    except Exception:
                        pass

                    try:
                        self.lbl_status.setText(f"SUP — filtrage spatial : {nom}...")
                        QApplication.processEvents()
                        if self.chk_intersect_only.isChecked():
                            res = processing.run("native:extractbylocation", {
                                'INPUT':     lyr_to_clip,
                                'PREDICATE': [0],
                                'INTERSECT': lyr_clip_effectif,
                                'OUTPUT':    'memory:'
                            })
                        else:
                            res = processing.run("native:clip", {
                                'INPUT':   lyr_to_clip,
                                'OVERLAY': lyr_clip_effectif,
                                'OUTPUT':  'memory:'
                            })
                        lyr_sup = res['OUTPUT']
                        lyr_sup.setName(nom)
                        if layer_has_features(lyr_sup):
                            lyr = lyr_sup
                    except Exception as e:
                        self.lbl_status.setText(f"⚠ Filtrage SUP échoué : {str(e)[:60]}")
                        QApplication.processEvents()

                if not layer_has_features(lyr):
                    self.lbl_status.setText(f"{nom} : aucune donnée dans la zone")
                    layers_empty.append(nom)
                    self.progress.setValue(i + 1)
                    QApplication.processEvents()
                    continue

                if self.chk_style.isChecked():
                    apply_style(lyr, fill, stroke)

                project.addMapLayer(lyr, False)
                if grp:
                    grp.addChildNode(QgsLayerTreeLayer(lyr))
                else:
                    root.addLayer(lyr)

                layers_loaded.append((nom, lyr))
                self.progress.setValue(i + 1)
                continue

            # ── Étape 1 : clip spatial rapide ─────────────────────────────
            # On évite fixgeometries par défaut : très coûteux sur les WFS.
            # On ne l'utilise qu'en secours si le clip direct échoue.
            if self.chk_clip.isChecked() and lyr_clip_effectif:
                try:
                    if self.chk_intersect_only.isChecked():
                        res = processing.run("native:extractbylocation", {
                            'INPUT':     lyr,
                            'PREDICATE': [0],
                            'INTERSECT': lyr_clip_effectif,
                            'OUTPUT':    'memory:'
                        })
                    else:
                        res = processing.run("native:clip", {
                            'INPUT':   lyr,
                            'OVERLAY': lyr_clip_effectif,
                            'OUTPUT':  'memory:'
                        })

                    lyr_clip = res['OUTPUT']
                    lyr_clip.setName(nom)
                    lyr = lyr_clip

                except Exception:
                    try:
                        lyr_fixed = processing.run("native:fixgeometries", {
                            'INPUT':  lyr,
                            'OUTPUT': 'memory:'
                        })['OUTPUT']

                        if self.chk_intersect_only.isChecked():
                            res = processing.run("native:extractbylocation", {
                                'INPUT':     lyr_fixed,
                                'PREDICATE': [0],
                                'INTERSECT': lyr_clip_effectif,
                                'OUTPUT':    'memory:'
                            })
                        else:
                            res = processing.run("native:clip", {
                                'INPUT':   lyr_fixed,
                                'OVERLAY': lyr_clip_effectif,
                                'OUTPUT':  'memory:'
                            })

                        lyr_clip = res['OUTPUT']
                        lyr_clip.setName(nom)
                        lyr = lyr_clip

                    except Exception as e:
                        self.lbl_status.setText(f"⚠ Clip échoué : {str(e)[:60]}")
                        QApplication.processEvents()

            # ── Étape 2 : filtre attributaire APRÈS le clip ───────────────
            if filtre_expr and layer_has_features(lyr):
                from qgis.core import QgsFeatureRequest, QgsExpression
                expr    = QgsExpression(filtre_expr)
                request = QgsFeatureRequest(expr)
                feats   = list(lyr.getFeatures(request))
                crs_id  = lyr.crs().authid()
                fields  = lyr.fields().toList()
                lyr_mem = QgsVectorLayer(
                    f"Polygon?crs={crs_id}", nom, "memory")
                pr_mem  = lyr_mem.dataProvider()
                pr_mem.addAttributes(fields)
                lyr_mem.updateFields()
                pr_mem.addFeatures(feats)
                lyr_mem.updateExtents()
                lyr = lyr_mem

            if not layer_has_features(lyr):
                self.lbl_status.setText(f"{nom} : aucune donnée après filtrage")
                layers_empty.append(nom)
                self.progress.setValue(i + 1)
                QApplication.processEvents()
                continue

            if self.chk_style.isChecked():
                apply_style(lyr, fill, stroke)

            project.addMapLayer(lyr, False)
            if grp:
                grp.addChildNode(QgsLayerTreeLayer(lyr))
            else:
                root.addLayer(lyr)

            layers_loaded.append((nom, lyr))
            self.progress.setValue(i + 1)

        if extent_effectif:
            ext = extent_effectif
            mx  = ext.width()  * 0.1
            my  = ext.height() * 0.1
            iface.mapCanvas().setExtent(
                type(ext)(ext.xMinimum()-mx, ext.yMinimum()-my,
                          ext.xMaximum()+mx, ext.yMaximum()+my))
            iface.mapCanvas().refresh()

        # Chargement GBIF si coché
        if self.chk_gbif.isChecked():
            self._load_gbif(project, root, grp,
                            extent_effectif, lyr_clip_effectif)

        if urbanisme_selected and not urbanisme_loaded:
            self._msgbox("info", "Données d'urbanisme",
                "La zone sélectionnée n'a pas de donnée d'urbanisme disponible.")


        nb_loaded = len(layers_loaded)
        self.lbl_status.setText(f"{nb_loaded} couche(s) chargee(s) avec succes !")
        self.lbl_status.setStyleSheet("color:#2A7A2A; font-weight:bold;")

        if layers_empty:
            self._msgbox(
                "info",
                "Couches sans données",
                "Les couches suivantes ne contiennent aucune donnée sur la zone sélectionnée "
                "et n'ont pas été ajoutées au projet :\n\n"
                + "\n".join(f"• {nom}" for nom in layers_empty)
            )

        self.btn_load.setEnabled(True)
         # Réinitialiser pour permettre un nouveau chargement
        # sur un territoire différent sans résidu de la sélection précédente
        self._territory_extent = None
        self._territory_lyr    = None
        self._territory_code   = None
        self._suggestions      = {}
        self.txt_search.clear()
        self._completer_model.setStringList([])
        self.lbl_result.setText("")
        self.progress.setVisible(False)

